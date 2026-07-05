import streamlit as st
import openpyxl
import pandas as pd
import os
import io
import tempfile
import sqlite3
import time
import re
from datetime import datetime
from PIL import Image as PILImage
from openpyxl.drawing.image import Image as OpenpyxlImage
import streamlit_authenticator as stauth
from sqlalchemy import create_engine
import boto3

# --- CONFIGURACIÓN DE LA NUBE ---
try:
    DB_URI = st.secrets["connections"]["postgresql"]["url"]
    R2_ENDPOINT = st.secrets["cloudflare"]["endpoint_url"]
    R2_ACCESS_KEY = st.secrets["cloudflare"]["access_key"]
    R2_SECRET_KEY = st.secrets["cloudflare"]["secret_key"]
    R2_BUCKET = st.secrets["cloudflare"]["bucket_name"]
    
    s3_client = boto3.client('s3', endpoint_url=R2_ENDPOINT, aws_access_key_id=R2_ACCESS_KEY, aws_secret_access_key=R2_SECRET_KEY)
    engine = create_engine(DB_URI)
except Exception as e:
    st.error("Error al cargar secretos de nube")
    st.stop()

st.set_page_config(page_title="Protocolos de Calidad | CCECC", page_icon="📝", layout="wide")


    



import base64

import base64

def get_base64_of_bin_file(bin_file):
    if os.path.exists(bin_file):
        with open(bin_file, 'rb') as f:
            data = f.read()
        return base64.b64encode(data).decode()
    return ""

logo_base64 = get_base64_of_bin_file("logo_china_civil.png")
fondo_base64 = get_base64_of_bin_file("fondo.png")

header_html = f"""
<style>
    /* Add the custom watermark background to the entire app */
    .stApp {{
        background-image: linear-gradient(rgba(255, 255, 255, 0.94), rgba(255, 255, 255, 0.94)), url("data:image/png;base64,{fondo_base64}");
        background-size: 80%;
        background-position: center center;
        background-repeat: no-repeat;
        background-attachment: fixed;
    }}
    
    /* Hide the default Streamlit header (with the Deploy button) */
    [data-testid="stHeader"] {{
        display: none;
    }}
    
    /* Add padding to the top of the main container so content isn't hidden behind our fixed header */
    .block-container {{
        padding-top: 7rem !important;
    }}
    
    /* Our custom sticky header */
    .custom-header {{
        position: fixed;
        top: 0;
        left: 0;
        right: 0;
        height: 85px;
        background: rgba(255, 255, 255, 0.65);
        backdrop-filter: blur(12px);
        -webkit-backdrop-filter: blur(12px);
        border-bottom: 1px solid rgba(255, 255, 255, 0.4);
        box-shadow: 0 8px 32px 0 rgba(31, 38, 135, 0.1);
        display: flex;
        align-items: center;
        justify-content: space-between;
        padding: 0 3rem;
        z-index: 999999;
    }}
    
    .header-logo {{
        display: flex;
        align-items: center;
        height: 100%;
    }}
    
    .header-text {{
        text-align: right;
    }}
    
    .header-title {{
        color: #1A365D;
        font-size: 1.3rem;
        font-weight: 700;
        margin: 0;
        letter-spacing: 0.5px;
    }}
    
    .header-subtitle {{
        font-size: 0.85rem;
        font-weight: 500;
        color: #555;
        margin: 0;
    }}
    /* Responsividad para Celulares */
    @media (max-width: 768px) {{
        .stApp {{
            background-attachment: scroll !important;
            background-size: cover !important;
            background-position: center top !important;
        }}
        .custom-header {{
            flex-direction: column;
            justify-content: center;
            height: auto;
            padding: 10px;
        }}
        .header-logo img {{
            height: 50px !important;
            margin-bottom: 5px;
        }}
        .header-text {{
            text-align: center;
        }}
        .header-title {{
            font-size: 1.1rem;
        }}
        .header-subtitle {{
            font-size: 0.75rem;
        }}
        .block-container {{
            padding-top: 9rem !important; /* Más espacio porque el header se apila */
        }}
    }}
</style>

<div class="custom-header">
    <div class="header-logo">
        <img src="data:image/png;base64,{logo_base64}" style="height: 75px; width: auto; object-fit: contain; margin-left: 5px;" />
    </div>
    <div class="header-text">
        <div class="header-title">SISTEMA DE GESTIÓN DE PROTOCOLOS DE CALIDAD</div>
        <div class="header-subtitle">Proyecto: Construcción de Infraestructura de Geología - Fase 6A y 6B</div>
    </div>
</div>
"""
st.markdown(header_html, unsafe_allow_html=True)

DB_FILE = "protocolos.db"
TEMPLATE_IPT = "LP16255BE-0138-0430-PTC-IPT-Inspeccion Pozo.xlsx"

# Helper to initialize DB
def load_db():
    try:
        with engine.connect() as conn:
            df = pd.read_sql("SELECT * FROM protocolos", conn)
        return df
    except Exception as e:
        return pd.DataFrame(columns=[
            "Correlativo", "Tipo_Protocolo", "Fecha", "Proyecto", "Tag", "Contratista", 
            "Tipo_Varilla", "Longitud", "Check1", "Check2", "Check3"
        ])

def generate_next_correlativo(current_val, tipo_prot):
    df_db = load_db()
    df_prot = df_db[df_db["Tipo_Protocolo"] == tipo_prot] if not df_db.empty else pd.DataFrame()
    
    max_num = 0
    if not df_prot.empty and "Correlativo" in df_prot.columns:
        for corr in df_prot["Correlativo"].dropna():
            match = re.search(r'-(\d+)$', str(corr))
            if match:
                num = int(match.group(1))
                if num > max_num:
                    max_num = num
    
    next_num = max_num + 1
    
    # Formato general de correlativos
    if tipo_prot == "PTC-RME":
        return f"LP16255BE-0840-0400-{tipo_prot}-{next_num:05d}"
    return f"LP16255BE-0840-0430-{tipo_prot}-{next_num:05d}"

def formato_titulo_es(texto):
    if not isinstance(texto, str) or not texto.strip(): return texto
    conectores = {"de", "del", "en", "para", "con", "por", "a", "sin", "sobre", "el", "la", "los", "las", "un", "una", "unos", "unas", "y", "o", "e", "u", "al"}
    palabras = texto.strip().split()
    if not palabras: return texto
    
    resultado = []
    for i, p in enumerate(palabras):
        # Si tiene números o ya está todo en mayúsculas (Ej. códigos, tags), no lo alteramos
        if any(c.isdigit() for c in p) or (p.isupper() and len(p) > 1):
            resultado.append(p)
        else:
            p_lower = p.lower()
            if i == 0 or p_lower not in conectores:
                resultado.append(p_lower.capitalize())
            else:
                resultado.append(p_lower)
    return " ".join(resultado)

def save_to_db(data_dict):
    exclude_keys = {"Correlativo", "Tipo_Protocolo", "Fecha", "Foto1", "Foto2", "Foto3", "Materiales_Data"}
    for key, value in data_dict.items():
        if key not in exclude_keys and isinstance(value, str):
            data_dict[key] = formato_titulo_es(value)

    tag = data_dict.get("Tag", "").strip() if data_dict.get("Tag") else ""
    fecha = data_dict.get("Fecha", "")
    tipo = data_dict.get("Tipo_Protocolo", "")
    
    if not tag or not fecha:
        raise ValueError("CAMPOS_VACIOS")

    try:
        with engine.begin() as conn:
            df = pd.read_sql("SELECT * FROM protocolos", conn)
            
            nuevo_correlativo = ""
            max_num = 0
            df_prot = df[df["Tipo_Protocolo"] == tipo] if (not df.empty and "Tipo_Protocolo" in df.columns) else pd.DataFrame()
            
            if not df_prot.empty and "Correlativo" in df_prot.columns:
                for corr in df_prot["Correlativo"].dropna():
                    match = re.search(r'-(\d+)$', str(corr))
                    if match:
                        num = int(match.group(1))
                        if num > max_num:
                            max_num = num
            
            next_num = max_num + 1
            if tipo == "PTC-RME":
                nuevo_correlativo = f"LP16255BE-0840-0400-{tipo}-{next_num:05d}"
            else:
                nuevo_correlativo = f"LP16255BE-0840-0430-{tipo}-{next_num:05d}"
                
            data_dict["Correlativo"] = nuevo_correlativo
            
            df_new = pd.DataFrame([data_dict])
            df_new.to_sql("protocolos", conn, if_exists="append", index=False)
            
            return nuevo_correlativo
    except Exception as e:
        raise e

def save_photos(uploaded_files, correlativo):
    paths = []
    if not os.path.exists("fotos_protocolos"):
        os.makedirs("fotos_protocolos")
        
    for i, file in enumerate(uploaded_files[:3]):
        timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
        filename = f"{correlativo}_foto{i+1}_{timestamp}.jpg"
        filepath = os.path.join("fotos_protocolos", filename)
        
        with open(filepath, "wb") as f:
            f.write(file.read())
        paths.append(filepath)
        
    while len(paths) < 3:
        paths.append("")
    return paths[0], paths[1], paths[2]

def render_check(text, options, key):
    col_txt, col_box = st.columns([3, 2])
    with col_txt:
        st.markdown(f"<div style='margin-top:7px; font-size:14.5px;'>{text}</div>", unsafe_allow_html=True)
    with col_box:
        return st.selectbox("##", options, key=key, label_visibility="collapsed")

# Estilos CSS
st.markdown("""
    <style>
    .stButton>button {
        height: 50px; font-size: 18px; border-radius: 10px;
        background-color: #004aad; color: white;
    }
    .stDownloadButton>button {
        background-color: #28a745; color: white;
    }
    </style>
""", unsafe_allow_html=True)


tab1, tab2, tab3 = st.tabs(["📝 Nuevo Registro (Campo)", "🗄️ Base de Datos e Impresión", "📊 Dashboard (Gerencia)"])

with tab1:
    st.markdown("Llene los datos desde campo para guardarlos en la base de datos oficial.")
    protocolo_seleccionado = st.selectbox(
        "Seleccione el Formato", 
        [
            "Inspección de Pozo a Tierra (PTC-IPT)", 
            "Resistencia de Pozo o Malla (PTC-MRP)",
            "Inspección Sistema de Puesta a Tierra (PTC-IST)",
            "Tratamiento de Malla (PTC-RTM)",
            "Tendido de Cable de Puesta a Tierra (PTC-TCA)",
            "Medición de Resistividad del Terreno (PTC-MRT)",
            "Recepción Materiales (PTC-RME)"
        ]
    )    
    # Extraer los últimos datos usados para pre-llenar los campos repetitivos
    df_defaults = load_db()
    if not df_defaults.empty:
        last_row = df_defaults.iloc[-1]
        def_proyecto = str(last_row.get("Proyecto", ""))
        def_cliente = str(last_row.get("Cliente", ""))
        def_contratista = str(last_row.get("Contratista", ""))
        def_contrato = str(last_row.get("Contrato", ""))
        def_disciplina = str(last_row.get("Disciplina", ""))
        def_equipo = str(last_row.get("Equipo_Prueba", ""))
        def_modelo = str(last_row.get("Modelo", ""))
        def_serie = str(last_row.get("N_Serie", ""))
    else:
        def_proyecto = "CONSTRUCCIÓN DE INFRAESTRUCTURA DE LAS NUEVAS FACILIDADES DE GEOLOGÍA - FASE 6A Y 6B"
        def_cliente = "MINERA LAS BAMBAS S.A."
        def_contratista = "CCECC"
        def_contrato = "CW2269622"
        def_disciplina = "0430 - Eléctrica"
        def_equipo = "Telurómetro Megger"
        def_modelo = "DET4TC2"
        def_serie = "SN-102938"
        
    # Variables de compatibilidad para los nuevos formatos
    default_proyecto = def_proyecto
    default_cliente = def_cliente
    default_contratista = def_contratista
    default_contrato = def_contrato
    default_lugar = ""
    default_facility = "0840 - Geology Facilities"
    
    if protocolo_seleccionado == "Inspección de Pozo a Tierra (PTC-IPT)":
        default_correlativo = generate_next_correlativo(None, "PTC-IPT")
        
        st.markdown(f"### 📋 PTC-IPT | Siguiente Correlativo: **{default_correlativo}**")
        with st.form("form_pozo", clear_on_submit=True):
            st.subheader("📋 1. Datos Generales")
            correlativo_input = st.text_input("Protocolo N° (Autogenerado / Editable)", value=default_correlativo)
            col1, col2, col3 = st.columns(3)
            with col1:
                fecha = st.date_input("Fecha de Ejecución")
                proyecto = st.text_input("Proyecto", value=def_proyecto)
                cliente = st.text_input("Cliente", value=def_cliente)
                lugar = st.text_input("Lugar/Área")
                contratista = st.text_input("Contratista", value=def_contratista)
            with col2:
                contrato = st.text_input("Contrato N°", value=def_contrato)
                plano = st.text_input("Plano de Referencia")
                facility = st.text_input("Facility code", value="0840 - Geology Facilities")
                tag = st.text_input("Tag del equipo")
                disciplina = st.text_input("Disciplina", value=def_disciplina)
            with col3:
                sistema = st.text_input("Sistema", value="No Aplica")
                subsistema = st.text_input("Sub sistema", value="No Aplica")
                estructura = st.text_input("Estructura")
                hoja = st.text_input("Hoja", value="1 de 3")
            
            st.subheader("🛠️ 2. Datos de Materiales")
            col_m1, col_m2, col_m3 = st.columns(3)
            with col_m1:
                tipo_varilla = st.text_input("Tipo de Varilla Vertical", value="COBRE ELECTROLÍTICO")
            with col_m2:
                seccion_varilla = st.text_input("Sección de Varilla Vertical", value='3/4"')
            with col_m3:
                longitud_varilla = st.text_input("Longitud de Varilla (m)", value="2.40 MT")
                
            col_c1, col_c2 = st.columns(2)
            with col_c1:
                conector = st.text_input("Conector Varilla Cable", value="GAR 6429")
            with col_c2:
                caja_reg = st.text_input("Caja de Registro de Puesta a Tierra", value="CAJA DE CONCRETO 400 X 400 X 300 mm")
            
            st.subheader("✅ 3. Checklist de Inspección")
            opciones_check = ["Conforme (C)", "No Conforme (NC)", "No Aplica (NA)"]
            check1 = render_check("1. Verificación Topográfica de ubicación del pozo a tierra.", opciones_check, "ipt_c1")
            check2 = render_check("2. Verificación de fundación de tratamiento del pozo a tierra está de acuerdo a planos y especificaciones.", opciones_check, "ipt_c2")
            check3 = render_check("3. Tratamiento de la puesta a tierra está de acuerdo a planos y especificaciones.", opciones_check, "ipt_c3")
            check4 = render_check("4. Verificación de relleno y compactación de puesta a tierra con material de acuerdo a planos y especificaciones.", opciones_check, "ipt_c4")
            check5 = render_check("5. Medición de resistencia puesta a tierra:", opciones_check, "ipt_c5")
            check6 = render_check("6. Ubicación e instalación de caja de registro de acuerdo a plano y especificaciones.", opciones_check, "ipt_c6")
            check7 = render_check("7. Conexión de ánodos de aterramiento revisados correctamente.", opciones_check, "ipt_c7")
            check8 = render_check("8. Rotulación y señalización de pozo a tierra.", opciones_check, "ipt_c8")
            
            st.subheader("💬 4. Comentarios y Observaciones")
            comentarios = st.text_area("Ingrese observaciones adicionales (opcional)")
            
            st.subheader("📸 5. Reporte Fotográfico (Opcional)")
            fotos_cargadas = st.file_uploader("📂 Sube fotos desde tu galería (hasta 3)", type=["jpg", "jpeg", "png"], accept_multiple_files=True, key="fotos_ipt_upload")
            
            st.markdown("---")
            submitted = st.form_submit_button("Guardar en Base de Datos 💾", use_container_width=True)
            
        if submitted:
            data = {
                "Correlativo": correlativo_input,
                "Tipo_Protocolo": "PTC-IPT",
                "Fecha": fecha.strftime("%d/%m/%Y"),
                "Proyecto": proyecto,
                "Cliente": cliente,
                "Lugar": lugar,
                "Contratista": contratista,
                "Contrato": contrato,
                "Plano": plano,
                "Facility": facility,
                "Tag": tag,
                "Disciplina": disciplina,
                "Sistema": sistema,
                "SubSistema": subsistema,
                "Estructura": estructura,
                "Hoja": hoja,
                "Tipo_Varilla": tipo_varilla,
                "Seccion_Varilla": seccion_varilla,
                "Longitud": longitud_varilla,
                "Conector": conector,
                "Caja_Reg": caja_reg,
                "Check1": check1,
                "Check2": check2,
                "Check3": check3,
                "Check4": check4,
                "Check5": check5,
                "Check6": check6,
                "Check7": check7,
                "Check8": check8,
                "Comentarios": comentarios
            }
            
            foto1, foto2, foto3 = save_photos(fotos_cargadas if fotos_cargadas else [], correlativo_input)
            data["Foto1"] = foto1
            data["Foto2"] = foto2
            data["Foto3"] = foto3
            
            try:
                correlativo = save_to_db(data)
                st.success(f"✅ ¡Registro guardado exitosamente con el número correlativo: **{correlativo}**!")
                st.info("Ve a la pestaña 'Base de Datos' para imprimir el documento.")
            except PermissionError:
                st.error("❌ ERROR: La base de datos está ocupada. Intenta de nuevo.")
            except ValueError as ve:
                if str(ve) == "CAMPOS_VACIOS":
                    st.markdown("""
                        <style>
                        div[data-baseweb="input"]:has(input[value=""]) {
                            border: 2px solid red !important;
                            border-radius: 4px;
                        }
                        </style>
                    """, unsafe_allow_html=True)
                    st.toast("⚠️ Faltan campos obligatorios. Revisa las celdas en rojo.", icon="❌")
                elif str(ve) == "DUPLICADO":
                    st.error("❌ ERROR: Este número correlativo ya existe. Por favor usa el siguiente número correlativo.")
                else:
                    st.error(f"❌ ERROR: {ve}")
            
    elif protocolo_seleccionado == "Resistencia de Pozo o Malla (PTC-MRP)":
        default_correlativo_mrp = generate_next_correlativo(None, "PTC-MRP")
        
        st.markdown(f"### 📋 PTC-MRP | Siguiente Correlativo: **{default_correlativo_mrp}**")
        with st.form("form_mrp", clear_on_submit=True):
            st.subheader("📋 1. Datos Generales")
            correlativo_input = st.text_input("Protocolo N° (Autogenerado / Editable)", value=default_correlativo_mrp)
            col1, col2, col3 = st.columns(3)
            with col1:
                fecha = st.date_input("Fecha de Ejecución")
                proyecto = st.text_input("Proyecto", value=def_proyecto)
                cliente = st.text_input("Cliente", value=def_cliente)
                lugar = st.text_input("Lugar / Área")
                contratista = st.text_input("Contratista", value=def_contratista)
            with col2:
                contrato = st.text_input("Contrato N°", value=def_contrato)
                plano = st.text_input("Plano de Referencia")
                facility = st.text_input("Facility code", value="0840 - Geology Facilities")
                tag = st.text_input("Tag del equipo")
                disciplina = st.text_input("Disciplina", value=def_disciplina)
            with col3:
                sistema = st.text_input("Sistema", value="No Aplica")
                subsistema = st.text_input("Sub Sistema", value="No Aplica")
                estructura = st.text_input("Estructura")
                hoja = st.text_input("Hoja", value="1 de 3")
            
            st.subheader("🛠️ 2. Equipos de Medición")
            col_eq1, col_eq2 = st.columns(2)
            with col_eq1:
                equipo_prueba = st.text_input("Equipo de Prueba", value="TELUROMETRO")
                marca = st.text_input("Marca", value="MEGABRAS")
            with col_eq2:
                modelo = st.text_input("Modelo", value="MTD20KWR")
                n_serie = st.text_input("Número de Serie", value="25L0108")
                fecha_calib = st.date_input("Fecha de Calibración", value=datetime.today())
                
            st.subheader("🌤️ 3. Datos Ambientales")
            col_amb1, col_amb2 = st.columns(2)
            with col_amb1:
                temperatura = st.text_input("Temperatura (°C)", value="12° C")
                humedad = st.text_input("Humedad Relativa (%)", value="45%")
            with col_amb2:
                lugar_prueba = st.text_input("Lugar de Prueba", value="Oficinas Geología 2")
                hora_prueba = st.time_input("Hora de Prueba", value=datetime.strptime("09:30", "%H:%M").time())
                
            st.subheader("⚡ 4. Mediciones (Caída de Potencial)")
            st.info("Ingresa la distancia R3 y los 5 valores medidos. Las distancias R2 se calcularán automáticamente en el Excel final.")
            dist_r3 = st.number_input("Distancia R3 (m)", value=14.5, format="%.1f")
            
            col_m1, col_m2, col_m3, col_m4, col_m5 = st.columns(5)
            with col_m1:
                val_42 = st.number_input("R a 42% (Ω)", value=10.60, format="%.2f")
            with col_m2:
                val_52 = st.number_input("R a 52% (Ω)", value=11.13, format="%.2f")
            with col_m3:
                val_62 = st.number_input("R a 62% (Ω)", value=11.70, format="%.2f")
            with col_m4:
                val_72 = st.number_input("R a 72% (Ω)", value=12.29, format="%.2f")
            with col_m5:
                val_82 = st.number_input("R a 82% (Ω)", value=13.62, format="%.2f")
                
            st.subheader("⚡ 5. Datos de la Malla / Conductor")
            col_cond1, col_cond2 = st.columns(2)
            with col_cond1:
                calibre = st.text_input("Calibre del Conductor", value="Varilla 3/4\"")
            with col_cond2:
                tipo_cond = st.text_input("Tipo de Conductor", value="Electrolítica")
            
            st.subheader("💬 6. Comentarios y Observaciones")
            comentarios = st.text_area("Ingrese observaciones adicionales (opcional)")
            
            st.subheader("📸 7. Reporte Fotográfico (Opcional)")
            fotos_cargadas_mrp = st.file_uploader("📂 Sube fotos desde tu galería (hasta 3)", type=["jpg", "jpeg", "png"], accept_multiple_files=True, key="fotos_mrp_upload")
            
            st.markdown("---")
            submitted_mrp = st.form_submit_button("💾 Guardar Protocolo en Base de Datos", use_container_width=True)
            
            if submitted_mrp:
                data = {
                    "Correlativo": correlativo_input,
                    "Tipo_Protocolo": "PTC-MRP",
                    "Fecha": fecha.strftime("%d/%m/%Y"),
                    "Proyecto": proyecto,
                    "Cliente": cliente,
                    "Lugar": lugar,
                    "Contratista": contratista,
                    "Contrato": contrato,
                    "Plano": plano,
                    "Facility": facility,
                    "Tag": tag,
                    "Disciplina": disciplina,
                    "Sistema": sistema,
                    "SubSistema": subsistema,
                    "Estructura": estructura,
                    "Hoja": hoja,
                    "Equipo_Prueba": equipo_prueba,
                    "Modelo": f"{marca} {modelo}",
                    "N_Serie": n_serie,
                    "Fecha_Calib": fecha_calib.strftime("%d/%m/%Y"),
                    "Temperatura": temperatura,
                    "Humedad": humedad,
                    "Lugar_Prueba": lugar_prueba,
                    "Hora_Prueba": hora_prueba.strftime("%H:%M"),
                    "Val_42": val_42,
                    "Val_52": val_52,
                    "Val_62": val_62,
                    "Val_72": val_72,
                    "Val_82": val_82,
                    "Dist_R3": dist_r3,
                    "Calibre": calibre,
                    "Tipo_Cond": tipo_cond,
                    "Comentarios": comentarios
                }
                foto1, foto2, foto3 = save_photos(fotos_cargadas_mrp if fotos_cargadas_mrp else [], correlativo_input)
                data["Foto1"] = foto1; data["Foto2"] = foto2; data["Foto3"] = foto3
                
                try:
                    save_to_db(data)
                    st.success(f"✅ ¡Guardado con éxito! {correlativo_input}")
                    st.rerun()
                except PermissionError:
                    st.error("❌ ERROR: El archivo 'BaseDatos_Protocolos.xlsx' está abierto en tu computadora. Por favor ciérralo y vuelve a presionar Guardar.")
                except ValueError as ve:
                    if str(ve) == "CAMPOS_VACIOS":
                        st.markdown("""
                            <style>
                            div[data-baseweb="input"]:has(input[value=""]) {
                                border: 2px solid red !important;
                                border-radius: 4px;
                            }
                            </style>
                        """, unsafe_allow_html=True)
                        st.toast("⚠️ Faltan campos obligatorios. Revisa las celdas en rojo.", icon="❌")
                    elif str(ve) == "DUPLICADO":
                        st.error("❌ ERROR: Este número correlativo ya existe. Por favor usa el siguiente número correlativo.")
                    else:
                        st.error(f"❌ ERROR: {ve}")

    elif protocolo_seleccionado == "Inspección Sistema de Puesta a Tierra (PTC-IST)":
        if "current_correlativo_ist" not in st.session_state:
            st.session_state.current_correlativo_ist = generate_next_correlativo("", "PTC-IST")
            
        st.markdown(f"### 📋 PTC-IST | Siguiente Correlativo: **{st.session_state.current_correlativo_ist}**")
        
        with st.form("form_ist", clear_on_submit=True):
            st.subheader("📝 1. Datos Generales")
            col1, col2, col3 = st.columns(3)
            with col1:
                fecha = st.date_input("Fecha de Ejecución", value=datetime.today())
                proyecto = st.text_input("Proyecto", value=default_proyecto)
                cliente = st.text_input("Cliente", value=default_cliente)
                lugar = st.text_input("Lugar/Área", value=default_lugar)
                plano = st.text_input("Plano de Referencia")
            with col2:
                contratista = st.text_input("Contratista", value=default_contratista)
                contrato = st.text_input("Contrato N°", value=default_contrato)
                facility = st.text_input("Facility code", value=default_facility)
                disciplina = st.text_input("Disciplina", value="0430 - Eléctrica")
            with col3:
                tag = st.text_input("Tag del Equipo", value="No Aplica")
                sistema = st.text_input("Sistema", value="No Aplica")
                subsistema = st.text_input("Sub sistema", value="No Aplica")
                estructura = st.text_input("Estructura", value="No Aplica")
                hoja = st.text_input("Hoja", value="1 de 4")
                
            st.subheader("✅ 2. Check List de Inspección")
            check_options = ["Conforme (C)", "No Conforme (NC)", "No Aplica (NA)"]
            checks_ist = []
            labels_ist = [
                "1. Verificacion Topografica de la malla del sistema PAT.",
                "2. El conductor primario (4/0 AWG) es del tamaño y tipo adecuado según los planos.",
                "3. Profundidad de enterramiento y recorrido del conductor de puesta a tierra según los planos y las especificaciones.",
                "4. El cable de tierra instalado cumple con especificación y planos.",
                "5. El cable de aterramiento se encuentra en buen estado.",
                "6. Los rellenos de la puesta a tierra se efectuaron correctamente.",
                "7. Las mechas cumplen con las distancias establecidas en los planos y están correctamente conectadas por compresión.",
                "8. Resistencia a tierra del ánodo del sistema de puesta a tierra comprobada y registrada.",
                "9. Se realizó la prueba de continuidad del lazo de aterramiento, correctamente.",
                "10. La resistencia a tierra máxima permitida no debe exceder 5 ohmios. Para los sistemas de puesta a tierra.",
                "11. Se utilizó tubería de PVC y abrazadera para subida de aterramiento.",
                "12. Se utilizó los elementos adecuados para el aterramiento (perno, tuerca, arandela plana y presión, terminal).",
                "13. Se utilizó terminal Burny, caña larga y caña corta, según especificaciones.",
                "14. Se conecto a presión el cable y terminal.",
                "15. Sellado de tubos (herméticos)."
            ]
            for i, label in enumerate(labels_ist):
                checks_ist.append(render_check(label, check_options, f"ist_c{i}"))
                
            st.subheader("💬 3. Comentarios y Observaciones")
            comentarios = st.text_area("Ingrese observaciones adicionales")
            
            st.subheader("📸 4. Reporte Fotográfico (Opcional)")
            fotos_cargadas_ist = st.file_uploader("📂 Sube fotos desde tu galería (hasta 3)", type=["jpg", "jpeg", "png"], accept_multiple_files=True, key="fotos_ist_upload")
            
            st.markdown("---")
            submitted_ist = st.form_submit_button("💾 Guardar Protocolo en Base de Datos", use_container_width=True)
            
            if submitted_ist:
                correlativo_input = st.session_state.current_correlativo_ist
                if len(fotos_cargadas_ist) > 3:
                    st.error("❌ Solo puedes subir hasta 3 fotos.")
                    st.stop()
                
                data = {
                    "Correlativo": correlativo_input,
                    "Tipo_Protocolo": "PTC-IST",
                    "Fecha": fecha.strftime("%d/%m/%Y"),
                    "Proyecto": proyecto, "Cliente": cliente, "Lugar": lugar, "Contratista": contratista,
                    "Contrato": contrato, "Plano": plano, "Facility": facility, "Tag": tag,
                    "Disciplina": disciplina, "Sistema": sistema, "SubSistema": subsistema,
                    "Estructura": estructura, "Hoja": hoja,
                    "Comentarios": comentarios
                }
                for i, check_val in enumerate(checks_ist):
                    data[f"Check{i+1}"] = check_val
                    
                foto1, foto2, foto3 = save_photos(fotos_cargadas_ist if fotos_cargadas_ist else [], correlativo_input)
                data["Foto1"] = foto1; data["Foto2"] = foto2; data["Foto3"] = foto3
                
                try:
                    correlativo = save_to_db(data)
                    st.session_state.current_correlativo_ist = generate_next_correlativo(correlativo, "PTC-IST")
                    st.success(f"✅ ¡Guardado con éxito! {correlativo}")
                    st.rerun()
                except PermissionError:
                    st.error("❌ ERROR: El archivo 'BaseDatos_Protocolos.xlsx' está abierto en tu computadora. Por favor ciérralo y vuelve a presionar Guardar.")
                except ValueError as ve:
                    if str(ve) == "CAMPOS_VACIOS":
                        st.markdown("""
                            <style>
                            div[data-baseweb="input"]:has(input[value=""]) {
                                border: 2px solid red !important;
                                border-radius: 4px;
                            }
                            </style>
                        """, unsafe_allow_html=True)
                        st.toast("⚠️ Faltan campos obligatorios. Revisa las celdas en rojo.", icon="❌")
                    elif str(ve) == "DUPLICADO":
                        st.error("❌ ERROR: Este número correlativo ya existe. Por favor usa el siguiente número correlativo.")
                    else:
                        st.error(f"❌ ERROR: {ve}")

    elif protocolo_seleccionado == "Tratamiento de Malla (PTC-RTM)":
        if "current_correlativo_rtm" not in st.session_state:
            st.session_state.current_correlativo_rtm = generate_next_correlativo("", "PTC-RTM")
            
        st.markdown(f"### 📋 PTC-RTM | Siguiente Correlativo: **{st.session_state.current_correlativo_rtm}**")
        
        with st.form("form_rtm", clear_on_submit=True):
            st.subheader("📝 1. Datos Generales")
            col1, col2, col3 = st.columns(3)
            with col1:
                fecha = st.date_input("Fecha de Ejecución", value=datetime.today())
                proyecto = st.text_input("Proyecto", value=default_proyecto)
                cliente = st.text_input("Cliente", value=default_cliente)
                lugar = st.text_input("Lugar/Área", value=default_lugar)
                plano = st.text_input("Plano de Referencia")
            with col2:
                contratista = st.text_input("Contratista", value=default_contratista)
                contrato = st.text_input("Contrato N°", value=default_contrato)
                facility = st.text_input("Facility code", value=default_facility)
                disciplina = st.text_input("Disciplina", value="0430 - Eléctrica")
            with col3:
                tag = st.text_input("Tag del Equipo", value="No Aplica")
                sistema = st.text_input("Sistema", value="No Aplica")
                subsistema = st.text_input("Sub sistema", value="No Aplica")
                estructura = st.text_input("Estructura", value="Malla a Tierra Principal")
                hoja = st.text_input("Hoja", value="1 de 3")
                
            st.subheader("🔌 2. Datos del Cable")
            col_c1, col_c2 = st.columns(2)
            with col_c1:
                fabricante = st.text_input("Fabricante", value="Celsa")
                seccion = st.text_input("Sección", value="120mm2")
                recorrido_desde = st.text_input("Recorrido Desde")
            with col_c2:
                longitud = st.text_input("Longitud", value="40 mts")
                tipo_cable = st.text_input("Tipo", value="Cobre desnudo")
                recorrido_hasta = st.text_input("Recorrido Hasta")
                
            st.subheader("✅ 3. Check List de Inspección")
            check_options = ["Conforme (C)", "No Conforme (NC)", "No Aplica (NA)"]
            checks_rtm = []
            labels_rtm = [
                "1. Verificación del trazo y replanteo del cable enterrado",
                "2. El topsoil fue cernido y está libre de material suelto y rocas",
                "3. La volumen del relleno preparado cumple con el 5% de bentonita y el resto de topsoil",
                "4. Se verificó la compactación de las 2 capas de relleno preparado",
                "5. Se verificó que el cable enterrado haya sido instalado sin deformaciones o separación de hebras",
                "6. Se inspeccionó el correcto prensado de las derivaciones de malla a tierra",
                "7. Se han usado los conectores correctos para los empalmes en las derivaciones",
                "8. Se verificó la dimensión transversal del cemento conductivo (Solo si aplicara)",
                "9. Se verificó el correcto fraguado del cemento conductivo antes de continuar con el relleno",
                "10. Se han dejado las reservas suficientes en derivaciones para las conexiones a equipos y estructuras",
                "11. Se ha cumplido con la profundidad mínima del cable de tierra (600 mm)",
                "12. Las actividades de excavación fueron revisadas y verificadas por personal civil especialista",
                "13. Se verificó la limpieza de la excavación (sin material suelto) y el nivel de fundacion fue verificado y liberado por la supervisión topográfica del cliente",
                "14. Verificación de la primera sección y segunda sección de relleno preparado (topsoil y bentonita)",
                "15. Se verifica la profundidad de instalación de la cinta de señalización indicada en la ingeniería del proyecto (300 mm)",
                "16. Se protegió el cable de tierra con tubería PVC en los cruces de estructuras de concreto"
            ]
            for i, label in enumerate(labels_rtm):
                checks_rtm.append(render_check(label, check_options, f"rtm_c{i}"))
                
            st.subheader("💬 4. Comentarios y Observaciones")
            comentarios = st.text_area("Ingrese observaciones adicionales")
            
            st.subheader("📸 5. Reporte Fotográfico (Opcional)")
            fotos_cargadas_rtm = st.file_uploader("📂 Sube fotos desde tu galería (hasta 3)", type=["jpg", "jpeg", "png"], accept_multiple_files=True, key="fotos_rtm_upload")
            
            st.markdown("---")
            submitted_rtm = st.form_submit_button("💾 Guardar Protocolo en Base de Datos", use_container_width=True)
            
            if submitted_rtm:
                correlativo_input = st.session_state.current_correlativo_rtm
                if len(fotos_cargadas_rtm) > 3:
                    st.error("❌ Solo puedes subir hasta 3 fotos.")
                    st.stop()
                
                data = {
                    "Correlativo": correlativo_input,
                    "Tipo_Protocolo": "PTC-RTM",
                    "Fecha": fecha.strftime("%d/%m/%Y"),
                    "Proyecto": proyecto, "Cliente": cliente, "Lugar": lugar, "Contratista": contratista,
                    "Contrato": contrato, "Plano": plano, "Facility": facility, "Tag": tag,
                    "Disciplina": disciplina, "Sistema": sistema, "SubSistema": subsistema,
                    "Estructura": estructura, "Hoja": hoja,
                    "Fabricante": fabricante, "Seccion": seccion, "Recorrido_Desde": recorrido_desde,
                    "Longitud": longitud, "Tipo_Cable": tipo_cable, "Recorrido_Hasta": recorrido_hasta,
                    "Comentarios": comentarios
                }
                for i, check_val in enumerate(checks_rtm):
                    data[f"Check{i+1}"] = check_val
                    
                foto1, foto2, foto3 = save_photos(fotos_cargadas_rtm if fotos_cargadas_rtm else [], correlativo_input)
                data["Foto1"] = foto1; data["Foto2"] = foto2; data["Foto3"] = foto3
                
                try:
                    correlativo = save_to_db(data)
                    st.session_state.current_correlativo_rtm = generate_next_correlativo(correlativo, "PTC-RTM")
                    st.success(f"✅ ¡Guardado con éxito! {correlativo}")
                    st.rerun()
                except PermissionError:
                    st.error("❌ ERROR: El archivo 'BaseDatos_Protocolos.xlsx' está abierto en tu computadora. Por favor ciérralo y vuelve a presionar Guardar.")
                except ValueError as ve:
                    if str(ve) == "CAMPOS_VACIOS":
                        st.markdown("""
                            <style>
                            div[data-baseweb="input"]:has(input[value=""]) {
                                border: 2px solid red !important;
                                border-radius: 4px;
                            }
                            </style>
                        """, unsafe_allow_html=True)
                        st.toast("⚠️ Faltan campos obligatorios. Revisa las celdas en rojo.", icon="❌")
                    elif str(ve) == "DUPLICADO":
                        st.error("❌ ERROR: Este número correlativo ya existe. Por favor usa el siguiente número correlativo.")
                    else:
                        st.error(f"❌ ERROR: {ve}")

    elif protocolo_seleccionado == "Tendido de Cable de Puesta a Tierra (PTC-TCA)":
        if "current_correlativo_tca" not in st.session_state:
            st.session_state.current_correlativo_tca = generate_next_correlativo("", "PTC-TCA")
            
        st.markdown(f"### 📋 PTC-TCA | Siguiente Correlativo: **{st.session_state.current_correlativo_tca}**")
        
        with st.form("form_tca", clear_on_submit=True):
            st.subheader("📝 1. Datos Generales")
            col1, col2, col3 = st.columns(3)
            with col1:
                fecha = st.date_input("Fecha de Ejecución", value=datetime.today())
                proyecto = st.text_input("Proyecto", value=default_proyecto)
                cliente = st.text_input("Cliente", value=default_cliente)
                lugar = st.text_input("Lugar/Área", value=default_lugar)
                plano = st.text_input("Plano de Referencia")
            with col2:
                contratista = st.text_input("Contratista", value=default_contratista)
                contrato = st.text_input("Contrato N°", value=default_contrato)
                facility = st.text_input("Facility code", value=default_facility)
                disciplina = st.text_input("Disciplina", value="0430 - Eléctrica")
            with col3:
                tag = st.text_input("Tag del Equipo", value="No Aplica")
                sistema = st.text_input("Sistema", value="No Aplica")
                subsistema = st.text_input("Sub sistema", value="No Aplica")
                estructura = st.text_input("Estructura", value="Malla de Tierra de Bancoductos")
                hoja = st.text_input("Hoja", value="1 de 3")
                
            st.subheader("🔌 2. Datos del Cable")
            col_c1, col_c2 = st.columns(2)
            with col_c1:
                fabricante = st.text_input("Fabricante", value="Celsa")
                seccion = st.text_input("Sección", value="4/0 AWG")
                recorrido_desde = st.text_input("Recorrido Desde")
            with col_c2:
                longitud = st.text_input("Longitud Total", value="31.8 mts")
                tipo_cable = st.text_input("Tipo", value="Cobre Desnudo")
                recorrido_hasta = st.text_input("Recorrido Hasta")
                
            st.subheader("🔗 3. Datos del Circuito (Tramos)")
            col_t1, col_t2, col_t3, col_t4 = st.columns(4)
            with col_t1: tag_tramo1 = st.text_input("Tag Tramo 1", value="No Aplica")
            with col_t2: desde_tramo1 = st.text_input("Desde Tramo 1", value="0840-HHA-0001")
            with col_t3: hasta_tramo1 = st.text_input("Hasta Tramo 1", value="0840-HHA-0002")
            with col_t4: long_tramo1 = st.text_input("Longitud Tramo 1", value="6.6 mts")
            
            col_u1, col_u2, col_u3, col_u4 = st.columns(4)
            with col_u1: tag_tramo2 = st.text_input("Tag Tramo 2", value="No Aplica")
            with col_u2: desde_tramo2 = st.text_input("Desde Tramo 2", value="0840-HHA-0002")
            with col_u3: hasta_tramo2 = st.text_input("Hasta Tramo 2", value="0840-HHA-0003")
            with col_u4: long_tramo2 = st.text_input("Longitud Tramo 2", value="25.2 mts")
                
            st.subheader("✅ 4. Check List de Inspección")
            check_options = ["Conforme (C)", "No Conforme (NC)", "No Aplica (NA)"]
            checks_tca = []
            labels_tca = [
                "1. Se verifico el cumplimiento de: sección, tipo, clase, numero de hilos, nivel de tensión",
                "2. Verificar el estado físico del cable, asegurándose de que no haya cortes, abolladuras o desgastes visibles",
                "3. Comprobar el aislamiento del cable para detectar cualquier daño que pueda comprometer su rendimiento o seguridad. Si Aplica",
                "4. Inspeccionar las conexiones de los cables a la barra de tierra o terminales, asegurando que estén bien apretadas y libres de corrosión.",
                "5. Medir la resistencia de puesta a tierra para confirmar que se encuentra dentro de los límites especificados por las normativas.",
                "6. Revisar la ruta del cable, verificando que esté instalado según el recorrido establecido sin obstrucciones ni excesiva tensión.",
                "7. Comprobar el radio de curvatura del cable, asegurándose de que no se haya doblado en ángulos demasiado cerrados que puedan dañar su estructura.",
                "8. Revisar la integridad de las bandejas y soportes donde se encuentra el cable, asegurando que ofrezcan soporte adecuado y protección contra posibles daños.",
                "9. Confirmar que el cable de cobre esté libre de corrosión o signos de degradación, especialmente en áreas expuestas a la humedad o ambientes corrosivos.",
                "10. Verificar que los conectores de cobre estén correctamente instalados y no presenten fisuras o corrosión.",
                "11. Comprobar que las conexiones estén bien apretadas y que no haya signos de aflojamiento.",
                "12. Asegurarse de que todas las instalaciones de cables y conectores cumplan con las normativas y estándares del cliente."
            ]
            for i, label in enumerate(labels_tca):
                checks_tca.append(render_check(label, check_options, f"tca_c{i}"))
                
            st.subheader("💬 5. Comentarios y Observaciones")
            comentarios = st.text_area("Ingrese observaciones adicionales")
            
            st.subheader("📸 6. Reporte Fotográfico (Opcional)")
            fotos_cargadas_tca = st.file_uploader("📂 Sube fotos desde tu galería (hasta 3)", type=["jpg", "jpeg", "png"], accept_multiple_files=True, key="fotos_tca_upload")
            
            st.markdown("---")
            submitted_tca = st.form_submit_button("💾 Guardar Protocolo en Base de Datos", use_container_width=True)
            
            if submitted_tca:
                correlativo_input = st.session_state.current_correlativo_tca
                if len(fotos_cargadas_tca) > 3:
                    st.error("❌ Solo puedes subir hasta 3 fotos.")
                    st.stop()
                
                data = {
                    "Correlativo": correlativo_input,
                    "Tipo_Protocolo": "PTC-TCA",
                    "Fecha": fecha.strftime("%d/%m/%Y"),
                    "Proyecto": proyecto, "Cliente": cliente, "Lugar": lugar, "Contratista": contratista,
                    "Contrato": contrato, "Plano": plano, "Facility": facility, "Tag": tag,
                    "Disciplina": disciplina, "Sistema": sistema, "SubSistema": subsistema,
                    "Estructura": estructura, "Hoja": hoja,
                    "Fabricante": fabricante, "Seccion": seccion, "Recorrido_Desde": recorrido_desde,
                    "Longitud": longitud, "Tipo_Cable": tipo_cable, "Recorrido_Hasta": recorrido_hasta,
                    "Tag_Tramo1": tag_tramo1, "Desde_Tramo1": desde_tramo1, "Hasta_Tramo1": hasta_tramo1, "Long_Tramo1": long_tramo1,
                    "Tag_Tramo2": tag_tramo2, "Desde_Tramo2": desde_tramo2, "Hasta_Tramo2": hasta_tramo2, "Long_Tramo2": long_tramo2,
                    "Comentarios": comentarios
                }
                for i, check_val in enumerate(checks_tca):
                    data[f"Check{i+1}"] = check_val
                    
                foto1, foto2, foto3 = save_photos(fotos_cargadas_tca if fotos_cargadas_tca else [], correlativo_input)
                data["Foto1"] = foto1; data["Foto2"] = foto2; data["Foto3"] = foto3
                
                try:
                    correlativo = save_to_db(data)
                    st.session_state.current_correlativo_tca = generate_next_correlativo(correlativo, "PTC-TCA")
                    st.success(f"✅ ¡Guardado con éxito! {correlativo}")
                    st.rerun()
                except PermissionError:
                    st.error("❌ ERROR: El archivo 'BaseDatos_Protocolos.xlsx' está abierto en tu computadora. Por favor ciérralo y vuelve a presionar Guardar.")
                except ValueError as ve:
                    if str(ve) == "CAMPOS_VACIOS":
                        st.markdown("""
                            <style>
                            div[data-baseweb="input"]:has(input[value=""]) {
                                border: 2px solid red !important;
                                border-radius: 4px;
                            }
                            </style>
                        """, unsafe_allow_html=True)
                        st.toast("⚠️ Faltan campos obligatorios. Revisa las celdas en rojo.", icon="❌")
                    elif str(ve) == "DUPLICADO":
                        st.error("❌ ERROR: Este número correlativo ya existe. Por favor usa el siguiente número correlativo.")
                    else:
                        st.error(f"❌ ERROR: {ve}")

    elif protocolo_seleccionado == "Medición de Resistividad del Terreno (PTC-MRT)":
        if "current_correlativo_mrt" not in st.session_state:
            st.session_state.current_correlativo_mrt = generate_next_correlativo("", "PTC-MRT")
            
        st.markdown(f"### 📋 PTC-MRT | Siguiente Correlativo: **{st.session_state.current_correlativo_mrt}**")
        
        with st.form("form_mrt", clear_on_submit=True):
            st.subheader("📝 1. Datos Generales")
            col1, col2, col3 = st.columns(3)
            with col1:
                fecha = st.date_input("Fecha de Ejecución", value=datetime.today())
                proyecto = st.text_input("Proyecto", value=default_proyecto)
                cliente = st.text_input("Cliente", value=default_cliente)
                lugar = st.text_input("Lugar/Área", value=default_lugar)
                plano = st.text_input("Plano de Referencia")
            with col2:
                contratista = st.text_input("Contratista", value=default_contratista)
                contrato = st.text_input("Contrato N°", value=default_contrato)
                facility = st.text_input("Facility code", value=default_facility)
                disciplina = st.text_input("Disciplina", value="0430 - Eléctrica")
            with col3:
                tag = st.text_input("Tag del Equipo", value="No Aplica")
                sistema = st.text_input("Sistema", value="No Aplica")
                subsistema = st.text_input("Sub sistema", value="No Aplica")
                estructura = st.text_input("Estructura", value="No Aplica")
                hoja = st.text_input("Hoja", value="1 de 4")
                
            st.subheader("🛠️ 2. Equipo de Medición")
            col_e1, col_e2, col_e3, col_e4 = st.columns(4)
            with col_e1: equipo_inst = st.text_input("Instrumento", value="Telurómetro")
            with col_e2: equipo_mod = st.text_input("Fab./Modelo", value="Megabrass / MTD20KWR")
            with col_e3: equipo_serie = st.text_input("N° Serie", value="25L0108")
            with col_e4: equipo_calib = st.date_input("Fecha Calibración", value=datetime(2025, 12, 12))
            
            st.subheader("📊 3. Tabulación de Mediciones")
            st.markdown("**(Valores por defecto b=0.35, a=1,2,4,8)**")
            
            perfiles_data = []
            for perfil in [1, 2]:
                st.markdown(f"#### Perfil {perfil}")
                for a_val in [1, 2, 4, 8]:
                    c1, c2, c3, c4, c5 = st.columns(5)
                    with c1: b_val = st.text_input(f"P{perfil}-a{a_val} | b (mts)", value="0.35", key=f"b_{perfil}_{a_val}")
                    with c2: a_inp = st.text_input(f"P{perfil}-a{a_val} | a (mts)", value=str(a_val), key=f"a_{perfil}_{a_val}")
                    with c3: r_val = st.text_input(f"P{perfil}-a{a_val} | R (ohms)", key=f"r_{perfil}_{a_val}")
                    with c4: hora = st.time_input(f"P{perfil}-a{a_val} | Hora", value=datetime.strptime("10:00", "%H:%M").time(), key=f"h_{perfil}_{a_val}")
                    with c5: temp = st.text_input(f"P{perfil}-a{a_val} | Temp (°C)", value="10.5", key=f"t_{perfil}_{a_val}")
                    perfiles_data.append({
                        "b": b_val, "a": a_inp, "r": r_val, 
                        "hora": hora.strftime("%H:%M:%S"), "temp": temp
                    })
                
            st.subheader("💬 4. Comentarios y Observaciones")
            comentarios = st.text_area("Ingrese observaciones adicionales")
            
            st.subheader("📸 5. Reporte Fotográfico (Opcional)")
            fotos_cargadas_mrt = st.file_uploader("📂 Sube fotos desde tu galería (hasta 3)", type=["jpg", "jpeg", "png"], accept_multiple_files=True, key="fotos_mrt_upload")
            
            st.markdown("---")
            submitted_mrt = st.form_submit_button("💾 Guardar Protocolo en Base de Datos", use_container_width=True)
            
            if submitted_mrt:
                correlativo_input = st.session_state.current_correlativo_mrt
                if len(fotos_cargadas_mrt) > 3:
                    st.error("❌ Solo puedes subir hasta 3 fotos.")
                    st.stop()
                
                data = {
                    "Correlativo": correlativo_input,
                    "Tipo_Protocolo": "PTC-MRT",
                    "Fecha": fecha.strftime("%d/%m/%Y"),
                    "Proyecto": proyecto, "Cliente": cliente, "Lugar": lugar, "Contratista": contratista,
                    "Contrato": contrato, "Plano": plano, "Facility": facility, "Tag": tag,
                    "Disciplina": disciplina, "Sistema": sistema, "SubSistema": subsistema,
                    "Estructura": estructura, "Hoja": hoja,
                    "Equipo_Inst": equipo_inst, "Equipo_Mod": equipo_mod, 
                    "Equipo_Serie": equipo_serie, "Equipo_Calib": equipo_calib.strftime("%d/%m/%Y"),
                    "Comentarios": comentarios,
                    "Perfiles_Data": str(perfiles_data)  # Saved as stringified list
                }
                foto1, foto2, foto3 = save_photos(fotos_cargadas_mrt if fotos_cargadas_mrt else [], correlativo_input)
                data["Foto1"] = foto1; data["Foto2"] = foto2; data["Foto3"] = foto3
                
                try:
                    correlativo = save_to_db(data)
                    st.session_state.current_correlativo_mrt = generate_next_correlativo(correlativo, "PTC-MRT")
                    st.success(f"✅ ¡Guardado con éxito! {correlativo}")
                    st.rerun()
                except PermissionError:
                    st.error("❌ ERROR: El archivo 'BaseDatos_Protocolos.xlsx' está abierto en tu computadora. Por favor ciérralo y vuelve a presionar Guardar.")
                except ValueError as ve:
                    if str(ve) == "CAMPOS_VACIOS":
                        st.markdown("""
                            <style>
                            div[data-baseweb="input"]:has(input[value=""]) {
                                border: 2px solid red !important;
                                border-radius: 4px;
                            }
                            </style>
                        """, unsafe_allow_html=True)
                        st.toast("⚠️ Faltan campos obligatorios. Revisa las celdas en rojo.", icon="❌")
                    elif str(ve) == "DUPLICADO":
                        st.error("❌ ERROR: Este número correlativo ya existe. Por favor usa el siguiente número correlativo.")
                    else:
                        st.error(f"❌ ERROR: {ve}")

    elif protocolo_seleccionado == "Recepción Materiales (PTC-RME)":
        if "current_correlativo_rme" not in st.session_state:
            st.session_state.current_correlativo_rme = generate_next_correlativo("", "PTC-RME")
            
        st.markdown(f"### 📋 PTC-RME | Siguiente Correlativo: **{st.session_state.current_correlativo_rme}**")
        
        with st.form("form_rme", clear_on_submit=True):
            st.subheader("📝 1. Datos Generales")
            col1, col2, col3 = st.columns(3)
            with col1:
                fecha = st.date_input("Fecha de Ejecución", value=datetime.today())
                proyecto = st.text_input("Proyecto", value=default_proyecto)
                cliente = st.text_input("Cliente", value=default_cliente)
                lugar = st.text_input("Lugar/Área", value=default_lugar)
                plano = st.text_input("Plano de Referencia", value="No Aplica")
            with col2:
                contratista = st.text_input("Contratista", value=default_contratista)
                contrato = st.text_input("Contrato N°", value=default_contrato)
                facility = st.text_input("Facility code", value=default_facility)
                disciplina = st.text_input("Disciplina", value="0400 - General")
            with col3:
                tag = st.text_input("Tag del Equipo", value="No Aplica")
                sistema = st.text_input("Sistema", value="No Aplica")
                subsistema = st.text_input("Sub sistema", value="No Aplica")
                estructura = st.text_input("Estructura", value="No Aplica")
                hoja = st.text_input("Hoja", value="1 de 16")
                
            st.subheader("📦 2. Lista de Materiales y Equipos")
            st.markdown("Abre las tarjetas que necesites llenar (hasta 30 ítems):")
            
            materiales_data_list = []
            unidades_opts = ["und", "m", "m2", "m3", "kg", "gln", "lts", "rollo", "varilla", "bolsa", "caja", "par", "pza", "otro"]
            cert_opts = ["Aplica", "No Aplica", "Faltante"]
            
            for i in range(1, 31):
                with st.expander(f"📦 Material {i}", expanded=(i == 1)):
                    col_m1, col_m2 = st.columns(2)
                    desc = col_m1.text_input("Descripción", key=f"rme_desc_{i}")
                    prov = col_m2.text_input("Proveedor", key=f"rme_prov_{i}")
                    
                    col_m3, col_m4, col_m5 = st.columns([1, 1, 1])
                    cant = col_m3.number_input("Cantidad", min_value=0.0, step=1.0, format="%.2f", key=f"rme_cant_{i}")
                    unidad = col_m4.selectbox("Unidad", unidades_opts, key=f"rme_und_{i}")
                    cert = col_m5.selectbox("Certificado", cert_opts, key=f"rme_cert_{i}")
                    
                    col_m6, col_m7 = st.columns(2)
                    guia = col_m6.text_input("Guía Remisión", key=f"rme_guia_{i}")
                    colada = col_m7.text_input("N° Colada", key=f"rme_colada_{i}")
                    
                    materiales_data_list.append({
                        "Descripción": desc,
                        "Proveedor": prov,
                        "Cantidad": cant,
                        "Unidad": unidad,
                        "Guía Remisión": guia,
                        "N° Colada": colada,
                        "Certificado": cert
                    })
            
            st.subheader("✅ 3. Resultados y Estado")
            col_r1, col_r2 = st.columns(2)
            with col_r1:
                resultado = st.radio("Resultados", ["Cumple", "No Cumple", "Observado"], horizontal=True)
            with col_r2:
                estado_aplica = st.radio("Equipo/Material Aplica:", ["Preservación", "Almacenamiento"], index=1, horizontal=True)
                
            st.subheader("💬 4. Observaciones")
            comentarios = st.text_area("Ingrese observaciones adicionales")
            
            st.subheader("📸 5. Reporte Fotográfico (Opcional)")
            fotos_cargadas_rme = st.file_uploader("📂 Sube fotos desde tu galería (hasta 3)", type=["jpg", "jpeg", "png"], accept_multiple_files=True, key="fotos_rme_upload")
            
            st.markdown("---")
            submitted_rme = st.form_submit_button("💾 Guardar Protocolo en Base de Datos", use_container_width=True)
            
            if submitted_rme:
                correlativo_input = st.session_state.current_correlativo_rme
                if len(fotos_cargadas_rme) > 3:
                    st.error("❌ Solo puedes subir hasta 3 fotos.")
                    st.stop()
                    
                # filter empty
                materiales_list = [m for m in materiales_data_list if str(m["Descripción"]).strip()]
                
                data = {
                    "Correlativo": correlativo_input,
                    "Tipo_Protocolo": "PTC-RME",
                    "Fecha": fecha.strftime("%d/%m/%Y"),
                    "Proyecto": proyecto, "Cliente": cliente, "Lugar": lugar, "Contratista": contratista,
                    "Contrato": contrato, "Plano": plano, "Facility": facility, "Tag": tag,
                    "Disciplina": disciplina, "Sistema": sistema, "SubSistema": subsistema,
                    "Estructura": estructura, "Hoja": hoja,
                    "Materiales_Data": str(materiales_list),
                    "Resultado": resultado, "Estado_Aplica": estado_aplica,
                    "Comentarios": comentarios
                }
                foto1, foto2, foto3 = save_photos(fotos_cargadas_rme if fotos_cargadas_rme else [], correlativo_input)
                data["Foto1"] = foto1; data["Foto2"] = foto2; data["Foto3"] = foto3
                
                try:
                    correlativo = save_to_db(data)
                    st.session_state.current_correlativo_rme = generate_next_correlativo(correlativo, "PTC-RME")
                    st.success(f"✅ ¡Guardado con éxito! {correlativo}")
                    st.rerun()
                except PermissionError:
                    st.error("❌ ERROR: El archivo 'BaseDatos_Protocolos.xlsx' está abierto en tu computadora. Por favor ciérralo y vuelve a presionar Guardar.")
                except ValueError as ve:
                    if str(ve) == "CAMPOS_VACIOS":
                        st.markdown("""
                            <style>
                            div[data-baseweb="input"]:has(input[value=""]) {
                                border: 2px solid red !important;
                                border-radius: 4px;
                            }
                            </style>
                        """, unsafe_allow_html=True)
                        st.toast("⚠️ Faltan campos obligatorios. Revisa las celdas en rojo.", icon="❌")
                    elif str(ve) == "DUPLICADO":
                        st.error("❌ ERROR: Este número correlativo ya existe. Por favor usa el siguiente número correlativo.")
                    else:
                        st.error(f"❌ ERROR: {ve}")

with tab2:
    st.markdown("Historial de protocolos. Selecciona uno para generar su formato final.")
    df = load_db()
    
    if not df.empty:
        # Botón para descargar toda la base de datos a Excel
        output_db = io.BytesIO()
        with pd.ExcelWriter(output_db, engine='openpyxl') as writer:
            df.to_excel(writer, index=False)
        output_db.seek(0)
        st.download_button(label="💾 Descargar Base de Datos Completa (Excel)", data=output_db, file_name="BaseDatos_Protocolos_Actualizada.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        
        st.markdown("### 🔍 Filtros de Búsqueda")
        col_f1, col_f2 = st.columns(2)
        with col_f1:
            filtro_tipo = st.multiselect("Filtrar por Tipo de Protocolo", df["Tipo_Protocolo"].unique())
        with col_f2:
            texto_busqueda = st.text_input("🔍 Buscar por Número, Tag o Fecha:")
                
        df_mostrar = df.copy()
        if filtro_tipo:
            df_mostrar = df_mostrar[df_mostrar["Tipo_Protocolo"].isin(filtro_tipo)]
        if texto_busqueda:
            texto_busqueda = texto_busqueda.lower()
            # Filtrar si el texto está en el Correlativo, Tag o Fecha
            mask = (
                df_mostrar["Correlativo"].astype(str).str.lower().str.contains(texto_busqueda) |
                (df_mostrar["Tag"] if "Tag" in df_mostrar.columns else pd.Series([""] * len(df_mostrar), index=df_mostrar.index)).astype(str).str.lower().str.contains(texto_busqueda) |
                (df_mostrar["Fecha"] if "Fecha" in df_mostrar.columns else pd.Series([""] * len(df_mostrar), index=df_mostrar.index)).astype(str).str.lower().str.contains(texto_busqueda)
            )
            df_mostrar = df_mostrar[mask]
            
        st.dataframe(df_mostrar, use_container_width=True)
        
        st.markdown("---")
        col_gen, col_del = st.columns(2)
        
        with col_gen:
            st.markdown("### 📄 Generar Documento")
            if not df_mostrar.empty:
                opciones = df_mostrar["Correlativo"].iloc[::-1].tolist()
                registro_a_generar = st.selectbox("Seleccione el correlativo a generar (Más recientes primero):", opciones, key="gen")
            else:
                registro_a_generar = None
                
            if registro_a_generar and st.button("Generar Plantilla Excel 📄"):
                row = df[df["Correlativo"] == registro_a_generar].iloc[0]
                
                if row["Tipo_Protocolo"] == "PTC-IPT":
                    if not os.path.exists(TEMPLATE_IPT):
                        st.error("❌ No se encontró la plantilla.")
                    else:
                        try:
                            wb = openpyxl.load_workbook(TEMPLATE_IPT)
                            ws = wb.active
                            
                            def set_value(sheet, cell_coord, val):
                                cell = sheet[cell_coord]
                                if type(cell).__name__ == 'MergedCell':
                                    for merged_range in sheet.merged_cells.ranges:
                                        if cell_coord in merged_range:
                                            top_left = merged_range.coord.split(':')[0]
                                            sheet[top_left] = val
                                            return
                                else:
                                    cell.value = val

                            set_value(ws, 'L5', row.get("Correlativo", ""))
                            set_value(ws, 'D4', row.get("Proyecto", ""))
                            set_value(ws, 'D5', row.get("Cliente", ""))
                            set_value(ws, 'D6', row.get("Lugar", ""))
                            set_value(ws, 'D7', row.get("Fecha", ""))
                            set_value(ws, 'D8', row.get("Plano", ""))
                            set_value(ws, 'D9', row.get("Tag", ""))
                            set_value(ws, 'D10', row.get("Sistema", ""))
                            set_value(ws, 'D11', row.get("Estructura", ""))
                            set_value(ws, 'L6', row.get("Contratista", ""))
                            set_value(ws, 'L7', row.get("Contrato", ""))
                            set_value(ws, 'L8', row.get("Facility", ""))
                            set_value(ws, 'L9', row.get("Disciplina", ""))
                            set_value(ws, 'L10', row.get("SubSistema", ""))
                            set_value(ws, 'L11', row.get("Hoja", ""))
                            
                            set_value(ws, 'C14', row.get("Tipo_Varilla", ""))
                            set_value(ws, 'I14', row.get("Seccion_Varilla", ""))
                            set_value(ws, 'N14', row.get("Longitud", ""))
                            set_value(ws, 'D18', row.get("Conector", ""))
                            set_value(ws, 'L18', row.get("Caja_Reg", ""))
                            
                            def apply_check(r, status):
                                set_value(ws, f'J{r}', "X" if status == "Conforme (C)" else "")
                                set_value(ws, f'K{r}', "X" if status == "No Conforme (NC)" else "")
                                set_value(ws, f'L{r}', "X" if status == "No Aplica (NA)" else "")
                            
                            apply_check(23, row.get("Check1", ""))
                            apply_check(24, row.get("Check2", ""))
                            apply_check(25, row.get("Check3", ""))
                            apply_check(26, row.get("Check4", ""))
                            apply_check(27, row.get("Check5", ""))
                            apply_check(29, row.get("Check6", ""))
                            apply_check(30, row.get("Check7", ""))
                            apply_check(31, row.get("Check8", ""))
                            
                            set_value(ws, 'A36', row.get("Comentarios", ""))
                            
                            def resize_image_for_excel(img_path, max_width=400, max_height=300):
                                with PILImage.open(img_path) as img:
                                    img.thumbnail((max_width, max_height))
                                    if img.mode in ("RGBA", "P"):
                                        img = img.convert("RGB")
                                    img_byte_arr = io.BytesIO()
                                    img.save(img_byte_arr, format="JPEG")
                                    img_byte_arr.seek(0)
                                    return img_byte_arr

                            has_photos = any(isinstance(row.get(f"Foto{i}"), str) and row.get(f"Foto{i}") for i in (1,2,3))
                            if has_photos:
                                ws_fotos = wb.create_sheet(title="Reporte Fotográfico")
                                ws_fotos['A1'] = "REPORTE FOTOGRÁFICO"
                                ws_fotos['A3'] = f"Proyecto: {row.get('Proyecto', '')}"
                                ws_fotos['A4'] = f"Protocolo N°: {row.get('Correlativo', '')}"
                                
                                start_rows = [7, 25, 43]
                                for idx, i in enumerate((1,2,3)):
                                    foto_path = row.get(f"Foto{i}")
                                    if isinstance(foto_path, str) and foto_path and os.path.exists(foto_path):
                                        try:
                                            resized_path = resize_image_for_excel(foto_path)
                                            img = OpenpyxlImage(resized_path)
                                            ws_fotos.add_image(img, f"B{start_rows[idx]}")
                                        except Exception as e:
                                            st.warning(f"No se pudo cargar Foto{i}: {e}")
                            
                            if "1" in wb.sheetnames:
                                del wb["1"]
                                
                            output = io.BytesIO()
                            wb.save(output)
                            output.seek(0)
                            
                            st.success(f"✅ ¡Plantilla para {registro_a_generar} lista!")
                            st.download_button(
                                label="⬇️ Descargar Excel Oficial",
                                data=output,
                                file_name=f"{registro_a_generar}_{row['Tag']}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                use_container_width=True
                            )
                        except Exception as e:
                            st.error(f"Error procesando el Excel: {e}")

                elif row["Tipo_Protocolo"] == "PTC-IST":
                    TEMPLATE_IST = "LP16255BE-0840-0430-PTC-IST - Inspección SPT 20.05.26.xlsx"
                    if not os.path.exists(TEMPLATE_IST):
                        st.error("❌ No se encontró la plantilla PTC-IST.")
                    else:
                        try:
                            wb = openpyxl.load_workbook(TEMPLATE_IST)
                            ws = wb.active
                            
                            def set_value(sheet, cell_coord, val):
                                cell = sheet[cell_coord]
                                if type(cell).__name__ == 'MergedCell':
                                    for merged_range in sheet.merged_cells.ranges:
                                        if cell_coord in merged_range:
                                            top_left = merged_range.coord.split(':')[0]
                                            sheet[top_left] = val
                                            return
                                else:
                                    cell.value = val

                            # Datos Generales
                            set_value(ws, 'L5', row.get("Correlativo", ""))
                            set_value(ws, 'D4', row.get("Proyecto", ""))
                            set_value(ws, 'D5', row.get("Cliente", ""))
                            set_value(ws, 'D6', row.get("Lugar", ""))
                            set_value(ws, 'D7', row.get("Fecha", ""))
                            set_value(ws, 'D8', row.get("Plano", ""))
                            set_value(ws, 'D9', row.get("Tag", ""))
                            set_value(ws, 'D10', row.get("Sistema", ""))
                            set_value(ws, 'D11', row.get("Estructura", ""))
                            set_value(ws, 'L6', row.get("Contratista", ""))
                            set_value(ws, 'L7', row.get("Contrato", ""))
                            set_value(ws, 'L8', row.get("Facility", ""))
                            set_value(ws, 'L9', row.get("Disciplina", ""))
                            set_value(ws, 'L10', row.get("SubSistema", ""))
                            set_value(ws, 'L11', row.get("Hoja", ""))
                            
                            # Checklists
                            def apply_check(r, status):
                                set_value(ws, f'J{r}', "X" if status == "Conforme (C)" else "")
                                set_value(ws, f'K{r}', "X" if status == "No Conforme (NC)" else "")
                                set_value(ws, f'L{r}', "X" if status == "No Aplica (NA)" else "")
                                
                            apply_check(15, row.get("Check1", ""))
                            apply_check(16, row.get("Check2", ""))
                            apply_check(17, row.get("Check3", ""))
                            apply_check(18, row.get("Check4", ""))
                            apply_check(19, row.get("Check5", ""))
                            apply_check(20, row.get("Check6", ""))
                            apply_check(21, row.get("Check7", ""))
                            apply_check(22, row.get("Check8", ""))
                            apply_check(23, row.get("Check9", ""))
                            apply_check(24, row.get("Check10", ""))
                            apply_check(25, row.get("Check11", ""))
                            apply_check(26, row.get("Check12", ""))
                            apply_check(27, row.get("Check13", ""))
                            apply_check(28, row.get("Check14", ""))
                            apply_check(29, row.get("Check15", ""))
                            
                            # Comentarios
                            set_value(ws, 'A35', row.get("Comentarios", ""))
                            
                            # FOTOS
                            def resize_image_for_excel(img_path, max_width=400, max_height=300):
                                with PILImage.open(img_path) as img:
                                    img.thumbnail((max_width, max_height))
                                    if img.mode in ("RGBA", "P"):
                                        img = img.convert("RGB")
                                    img_byte_arr = io.BytesIO()
                                    img.save(img_byte_arr, format="JPEG")
                                    img_byte_arr.seek(0)
                                    return img_byte_arr
                            has_photos = any(isinstance(row.get(f"Foto{i}"), str) and row.get(f"Foto{i}") for i in (1,2,3))
                            if has_photos:
                                ws_fotos = wb.create_sheet(title="Reporte Fotográfico")
                                ws_fotos['A1'] = "REPORTE FOTOGRÁFICO"
                                ws_fotos['A3'] = f"Proyecto: {row.get('Proyecto', '')}"
                                ws_fotos['A4'] = f"Protocolo N°: {row.get('Correlativo', '')}"
                                start_rows = [7, 25, 43]
                                for idx, i in enumerate((1,2,3)):
                                    foto_path = row.get(f"Foto{i}")
                                    if isinstance(foto_path, str) and foto_path and os.path.exists(foto_path):
                                        try:
                                            resized_path = resize_image_for_excel(foto_path)
                                            img = OpenpyxlImage(resized_path)
                                            ws_fotos.add_image(img, f"B{start_rows[idx]}")
                                        except Exception as e:
                                            st.warning(f"No se pudo cargar Foto{i}: {e}")
                                            
                            if "1" in wb.sheetnames: del wb["1"]
                            output = io.BytesIO(); wb.save(output); output.seek(0)
                            
                            st.success(f"✅ ¡Plantilla para {registro_a_generar} lista!")
                            st.download_button("⬇️ Descargar Excel Oficial", data=output, file_name=f"{registro_a_generar}_{row['Tag']}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
                        except Exception as e:
                            st.error(f"Error procesando el Excel: {e}")
                            
                elif row["Tipo_Protocolo"] == "PTC-RTM":
                    TEMPLATE_RTM = "LP16255BE-0840-0430-PTC-RTM-Tratamiento de malla a tierra.xlsx"
                    if not os.path.exists(TEMPLATE_RTM):
                        st.error("❌ No se encontró la plantilla PTC-RTM.")
                    else:
                        try:
                            wb = openpyxl.load_workbook(TEMPLATE_RTM)
                            ws = wb.active
                            
                            def set_value(sheet, cell_coord, val):
                                cell = sheet[cell_coord]
                                if type(cell).__name__ == 'MergedCell':
                                    for merged_range in sheet.merged_cells.ranges:
                                        if cell_coord in merged_range:
                                            top_left = merged_range.coord.split(':')[0]
                                            sheet[top_left] = val
                                            return
                                else:
                                    cell.value = val

                            # Datos Generales
                            set_value(ws, 'M5', row.get("Correlativo", ""))
                            set_value(ws, 'D4', row.get("Proyecto", ""))
                            set_value(ws, 'D5', row.get("Cliente", ""))
                            set_value(ws, 'D6', row.get("Lugar", ""))
                            set_value(ws, 'D7', row.get("Fecha", ""))
                            set_value(ws, 'D8', row.get("Plano", ""))
                            set_value(ws, 'D9', row.get("Tag", ""))
                            set_value(ws, 'D10', row.get("Sistema", ""))
                            set_value(ws, 'D11', row.get("Estructura", ""))
                            set_value(ws, 'M6', row.get("Contratista", ""))
                            set_value(ws, 'M7', row.get("Contrato", ""))
                            set_value(ws, 'M8', row.get("Facility", ""))
                            set_value(ws, 'M9', row.get("Disciplina", ""))
                            set_value(ws, 'M10', row.get("SubSistema", ""))
                            set_value(ws, 'M11', row.get("Hoja", ""))
                            
                            # Datos del Cable
                            set_value(ws, 'D13', row.get("Fabricante", ""))
                            set_value(ws, 'M13', row.get("Longitud", ""))
                            set_value(ws, 'D14', row.get("Seccion", ""))
                            set_value(ws, 'M14', row.get("Tipo_Cable", ""))
                            set_value(ws, 'D15', row.get("Recorrido_Desde", ""))
                            set_value(ws, 'M15', row.get("Recorrido_Hasta", ""))
                            
                            # Checklists
                            def apply_check(r, status):
                                set_value(ws, f'J{r}', "X" if status == "Conforme (C)" else "")
                                set_value(ws, f'K{r}', "X" if status == "No Conforme (NC)" else "")
                                set_value(ws, f'L{r}', "X" if status == "No Aplica (NA)" else "")
                                
                            for i in range(1, 17):
                                apply_check(17 + i, row.get(f"Check{i}", ""))
                            
                            # Comentarios
                            set_value(ws, 'A37', row.get("Comentarios", ""))
                            
                            # FOTOS
                            def resize_image_for_excel(img_path, max_width=400, max_height=300):
                                with PILImage.open(img_path) as img:
                                    img.thumbnail((max_width, max_height))
                                    if img.mode in ("RGBA", "P"):
                                        img = img.convert("RGB")
                                    img_byte_arr = io.BytesIO()
                                    img.save(img_byte_arr, format="JPEG")
                                    img_byte_arr.seek(0)
                                    return img_byte_arr
                            has_photos = any(isinstance(row.get(f"Foto{i}"), str) and row.get(f"Foto{i}") for i in (1,2,3))
                            if has_photos:
                                ws_fotos = wb.create_sheet(title="Reporte Fotográfico")
                                ws_fotos['A1'] = "REPORTE FOTOGRÁFICO"
                                ws_fotos['A3'] = f"Proyecto: {row.get('Proyecto', '')}"
                                ws_fotos['A4'] = f"Protocolo N°: {row.get('Correlativo', '')}"
                                start_rows = [7, 25, 43]
                                for idx, i in enumerate((1,2,3)):
                                    foto_path = row.get(f"Foto{i}")
                                    if isinstance(foto_path, str) and foto_path and os.path.exists(foto_path):
                                        try:
                                            resized_path = resize_image_for_excel(foto_path)
                                            img = OpenpyxlImage(resized_path)
                                            ws_fotos.add_image(img, f"B{start_rows[idx]}")
                                        except Exception as e:
                                            st.warning(f"No se pudo cargar Foto{i}: {e}")
                                            
                            if "1" in wb.sheetnames: del wb["1"]
                            output = io.BytesIO(); wb.save(output); output.seek(0)
                            
                            st.success(f"✅ ¡Plantilla para {registro_a_generar} lista!")
                            st.download_button("⬇️ Descargar Excel Oficial", data=output, file_name=f"{registro_a_generar}_{row['Tag']}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
                        except Exception as e:
                            st.error(f"Error procesando el Excel: {e}")
                            
                elif row["Tipo_Protocolo"] == "PTC-TCA":
                    TEMPLATE_TCA = "LP16255BE-0840-0430-PTC-TCA-Tendido de Cable de Atierra.xlsx"
                    if not os.path.exists(TEMPLATE_TCA):
                        st.error("❌ No se encontró la plantilla PTC-TCA.")
                    else:
                        try:
                            wb = openpyxl.load_workbook(TEMPLATE_TCA)
                            ws = wb.active
                            
                            def set_value(sheet, cell_coord, val):
                                cell = sheet[cell_coord]
                                if type(cell).__name__ == 'MergedCell':
                                    for merged_range in sheet.merged_cells.ranges:
                                        if cell_coord in merged_range:
                                            top_left = merged_range.coord.split(':')[0]
                                            sheet[top_left] = val
                                            return
                                else:
                                    cell.value = val

                            # Datos Generales
                            set_value(ws, 'M5', row.get("Correlativo", ""))
                            set_value(ws, 'D4', row.get("Proyecto", ""))
                            set_value(ws, 'D5', row.get("Cliente", ""))
                            set_value(ws, 'D6', row.get("Lugar", ""))
                            set_value(ws, 'D7', row.get("Fecha", ""))
                            set_value(ws, 'D8', row.get("Plano", ""))
                            set_value(ws, 'D9', row.get("Tag", ""))
                            set_value(ws, 'D10', row.get("Sistema", ""))
                            set_value(ws, 'D11', row.get("Estructura", ""))
                            set_value(ws, 'M6', row.get("Contratista", ""))
                            set_value(ws, 'M7', row.get("Contrato", ""))
                            set_value(ws, 'M8', row.get("Facility", ""))
                            set_value(ws, 'M9', row.get("Disciplina", ""))
                            set_value(ws, 'M10', row.get("SubSistema", ""))
                            set_value(ws, 'M11', row.get("Hoja", ""))
                            
                            # Datos del Cable
                            set_value(ws, 'D13', row.get("Fabricante", ""))
                            set_value(ws, 'M13', row.get("Longitud", ""))
                            set_value(ws, 'D14', row.get("Seccion", ""))
                            set_value(ws, 'M14', row.get("Tipo_Cable", ""))
                            set_value(ws, 'D15', row.get("Recorrido_Desde", ""))
                            set_value(ws, 'M15', row.get("Recorrido_Hasta", ""))
                            
                            # Datos del Circuito
                            set_value(ws, 'A18', row.get("Tag_Tramo1", ""))
                            set_value(ws, 'E18', row.get("Desde_Tramo1", ""))
                            set_value(ws, 'I18', row.get("Hasta_Tramo1", ""))
                            set_value(ws, 'M18', row.get("Long_Tramo1", ""))
                            
                            set_value(ws, 'A19', row.get("Tag_Tramo2", ""))
                            set_value(ws, 'E19', row.get("Desde_Tramo2", ""))
                            set_value(ws, 'I19', row.get("Hasta_Tramo2", ""))
                            set_value(ws, 'M19', row.get("Long_Tramo2", ""))
                            
                            # Checklists
                            def apply_check(r, status):
                                set_value(ws, f'J{r}', "X" if status == "Conforme (C)" else "")
                                set_value(ws, f'K{r}', "X" if status == "No Conforme (NC)" else "")
                                set_value(ws, f'L{r}', "X" if status == "No Aplica (NA)" else "")
                                
                            for i in range(1, 13):
                                apply_check(22 + i, row.get(f"Check{i}", ""))
                            
                            # Comentarios
                            set_value(ws, 'A37', row.get("Comentarios", ""))
                            
                            # FOTOS
                            def resize_image_for_excel(img_path, max_width=400, max_height=300):
                                with PILImage.open(img_path) as img:
                                    img.thumbnail((max_width, max_height))
                                    if img.mode in ("RGBA", "P"):
                                        img = img.convert("RGB")
                                    img_byte_arr = io.BytesIO()
                                    img.save(img_byte_arr, format="JPEG")
                                    img_byte_arr.seek(0)
                                    return img_byte_arr
                            has_photos = any(isinstance(row.get(f"Foto{i}"), str) and row.get(f"Foto{i}") for i in (1,2,3))
                            if has_photos:
                                ws_fotos = wb.create_sheet(title="Reporte Fotográfico")
                                ws_fotos['A1'] = "REPORTE FOTOGRÁFICO"
                                ws_fotos['A3'] = f"Proyecto: {row.get('Proyecto', '')}"
                                ws_fotos['A4'] = f"Protocolo N°: {row.get('Correlativo', '')}"
                                start_rows = [7, 25, 43]
                                for idx, i in enumerate((1,2,3)):
                                    foto_path = row.get(f"Foto{i}")
                                    if isinstance(foto_path, str) and foto_path and os.path.exists(foto_path):
                                        try:
                                            resized_path = resize_image_for_excel(foto_path)
                                            img = OpenpyxlImage(resized_path)
                                            ws_fotos.add_image(img, f"B{start_rows[idx]}")
                                        except Exception as e:
                                            st.warning(f"No se pudo cargar Foto{i}: {e}")
                                            
                            if "1" in wb.sheetnames: del wb["1"]
                            output = io.BytesIO(); wb.save(output); output.seek(0)
                            
                            st.success(f"✅ ¡Plantilla para {registro_a_generar} lista!")
                            st.download_button("⬇️ Descargar Excel Oficial", data=output, file_name=f"{registro_a_generar}_{row['Tag']}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
                        except Exception as e:
                            st.error(f"Error procesando el Excel: {e}")
                            
                elif row["Tipo_Protocolo"] == "PTC-MRT":
                    TEMPLATE_MRT = "LP16255BE-0840-0430-PTC-MRT_24.04.26.xlsx"
                    if not os.path.exists(TEMPLATE_MRT):
                        st.error("❌ No se encontró la plantilla PTC-MRT.")
                    else:
                        try:
                            wb = openpyxl.load_workbook(TEMPLATE_MRT)
                            ws = wb.active
                            
                            def set_value(sheet, cell_coord, val):
                                cell = sheet[cell_coord]
                                if type(cell).__name__ == 'MergedCell':
                                    for merged_range in sheet.merged_cells.ranges:
                                        if cell_coord in merged_range:
                                            top_left = merged_range.coord.split(':')[0]
                                            sheet[top_left] = val
                                            return
                                else:
                                    cell.value = val

                            # Datos Generales
                            set_value(ws, 'L5', row.get("Correlativo", ""))
                            set_value(ws, 'D4', row.get("Proyecto", ""))
                            set_value(ws, 'D5', row.get("Cliente", ""))
                            set_value(ws, 'D6', row.get("Lugar", ""))
                            set_value(ws, 'D7', row.get("Fecha", ""))
                            set_value(ws, 'D8', row.get("Plano", ""))
                            set_value(ws, 'D9', row.get("Tag", ""))
                            set_value(ws, 'D10', row.get("Sistema", ""))
                            set_value(ws, 'D11', row.get("Estructura", ""))
                            set_value(ws, 'L6', row.get("Contratista", ""))
                            set_value(ws, 'L7', row.get("Contrato", ""))
                            set_value(ws, 'L8', row.get("Facility", ""))
                            set_value(ws, 'L9', row.get("Disciplina", ""))
                            set_value(ws, 'L10', row.get("SubSistema", ""))
                            set_value(ws, 'L11', row.get("Hoja", ""))
                            
                            # Equipo
                            set_value(ws, 'A14', row.get("Equipo_Inst", ""))
                            set_value(ws, 'E14', row.get("Equipo_Mod", ""))
                            set_value(ws, 'I14', row.get("Equipo_Serie", ""))
                            set_value(ws, 'M14', row.get("Equipo_Calib", ""))
                            
                            # Perfiles Data
                            import ast
                            try:
                                perfiles_data = ast.literal_eval(row.get("Perfiles_Data", "[]"))
                                start_r = 28
                                for idx, pdata in enumerate(perfiles_data):
                                    r = start_r + idx
                                    set_value(ws, f'C{r}', pdata.get("b", ""))
                                    set_value(ws, f'D{r}', pdata.get("a", ""))
                                    set_value(ws, f'F{r}', pdata.get("r", ""))
                                    set_value(ws, f'J{r}', pdata.get("hora", ""))
                                    set_value(ws, f'L{r}', pdata.get("temp", ""))
                            except:
                                pass
                            
                            # Comentarios
                            set_value(ws, 'A39', row.get("Comentarios", ""))
                            
                            # FOTOS
                            def resize_image_for_excel(img_path, max_width=400, max_height=300):
                                with PILImage.open(img_path) as img:
                                    img.thumbnail((max_width, max_height))
                                    if img.mode in ("RGBA", "P"):
                                        img = img.convert("RGB")
                                    img_byte_arr = io.BytesIO()
                                    img.save(img_byte_arr, format="JPEG")
                                    img_byte_arr.seek(0)
                                    return img_byte_arr
                            has_photos = any(isinstance(row.get(f"Foto{i}"), str) and row.get(f"Foto{i}") for i in (1,2,3))
                            if has_photos:
                                ws_fotos = wb.create_sheet(title="Reporte Fotográfico")
                                ws_fotos['A1'] = "REPORTE FOTOGRÁFICO"
                                ws_fotos['A3'] = f"Proyecto: {row.get('Proyecto', '')}"
                                ws_fotos['A4'] = f"Protocolo N°: {row.get('Correlativo', '')}"
                                start_rows = [7, 25, 43]
                                for idx, i in enumerate((1,2,3)):
                                    foto_path = row.get(f"Foto{i}")
                                    if isinstance(foto_path, str) and foto_path and os.path.exists(foto_path):
                                        try:
                                            resized_path = resize_image_for_excel(foto_path)
                                            img = OpenpyxlImage(resized_path)
                                            ws_fotos.add_image(img, f"B{start_rows[idx]}")
                                        except Exception as e:
                                            st.warning(f"No se pudo cargar Foto{i}: {e}")
                                            
                            if "1" in wb.sheetnames: del wb["1"]
                            output = io.BytesIO(); wb.save(output); output.seek(0)
                            
                            st.success(f"✅ ¡Plantilla para {registro_a_generar} lista!")
                            st.download_button("⬇️ Descargar Excel Oficial", data=output, file_name=f"{registro_a_generar}_{row['Tag']}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
                        except Exception as e:
                            st.error(f"Error procesando el Excel: {e}")
                            
                elif row["Tipo_Protocolo"] == "PTC-RME":
                    TEMPLATE_RME = "LP16255BE-0840-0400-PTC-RME_24.04.26.xlsx"
                    if not os.path.exists(TEMPLATE_RME):
                        st.error("❌ No se encontró la plantilla PTC-RME.")
                    else:
                        try:
                            wb = openpyxl.load_workbook(TEMPLATE_RME)
                            ws = wb.active
                            
                            def set_value(sheet, cell_coord, val):
                                cell = sheet[cell_coord]
                                if type(cell).__name__ == 'MergedCell':
                                    for merged_range in sheet.merged_cells.ranges:
                                        if cell_coord in merged_range:
                                            top_left = merged_range.coord.split(':')[0]
                                            sheet[top_left] = val
                                            return
                                else:
                                    cell.value = val

                            # Datos Generales
                            set_value(ws, 'L5', row.get("Correlativo", ""))
                            set_value(ws, 'D4', row.get("Proyecto", ""))
                            set_value(ws, 'D5', row.get("Cliente", ""))
                            set_value(ws, 'D6', row.get("Lugar", ""))
                            set_value(ws, 'D7', row.get("Fecha", ""))
                            set_value(ws, 'D8', row.get("Plano", ""))
                            set_value(ws, 'D9', row.get("Tag", ""))
                            set_value(ws, 'D10', row.get("Sistema", ""))
                            set_value(ws, 'D11', row.get("Estructura", ""))
                            set_value(ws, 'L6', row.get("Contratista", ""))
                            set_value(ws, 'L7', row.get("Contrato", ""))
                            set_value(ws, 'L8', row.get("Facility", ""))
                            set_value(ws, 'L9', row.get("Disciplina", ""))
                            set_value(ws, 'L10', row.get("SubSistema", ""))
                            set_value(ws, 'L11', row.get("Hoja", ""))
                            
                            # Materiales Data
                            import ast
                            try:
                                materiales = ast.literal_eval(row.get("Materiales_Data", "[]"))
                                start_r = 13
                                for idx, mdata in enumerate(materiales):
                                    r = start_r + idx
                                    set_value(ws, f'B{r}', str(mdata.get("Descripción", "")))
                                    set_value(ws, f'G{r}', str(mdata.get("Proveedor", "")))
                                    set_value(ws, f'I{r}', str(mdata.get("Cantidad", "")))
                                    set_value(ws, f'J{r}', str(mdata.get("Unidad", "")))
                                    set_value(ws, f'K{r}', str(mdata.get("Guía Remisión", "")))
                                    set_value(ws, f'L{r}', str(mdata.get("N° Colada", "")))
                                    set_value(ws, f'N{r}', str(mdata.get("Certificado", "")))
                            except Exception as em:
                                print(f"Error materiales: {em}")
                                pass
                                
                            # Resultados & Estado
                            res = row.get("Resultado", "")
                            ws['A37'] = f"RESULTADOS:                                     CUMPLE ({' X ' if res=='Cumple' else '   '})                 NO CUMPLE ({' X ' if res=='No Cumple' else '   '})                    OBSERVADO  ({' X ' if res=='Observado' else '   '})"
                            
                            apl = row.get("Estado_Aplica", "")
                            ws['A38'] = f"EQUIPO Y/ O MATERIAL APLICA:     PRESERVACIÓN ({' X ' if apl=='Preservación' else '   '})                 ALMACENAMIENTO    ({' X ' if apl=='Almacenamiento' else '   '})"
                            
                            # Comentarios
                            set_value(ws, 'D39', row.get("Comentarios", ""))
                            
                            # FOTOS
                            def resize_image_for_excel(img_path, max_width=400, max_height=300):
                                with PILImage.open(img_path) as img:
                                    img.thumbnail((max_width, max_height))
                                    if img.mode in ("RGBA", "P"):
                                        img = img.convert("RGB")
                                    img_byte_arr = io.BytesIO()
                                    img.save(img_byte_arr, format="JPEG")
                                    img_byte_arr.seek(0)
                                    return img_byte_arr
                            has_photos = any(isinstance(row.get(f"Foto{i}"), str) and row.get(f"Foto{i}") for i in (1,2,3))
                            if has_photos:
                                ws_fotos = wb.create_sheet(title="Reporte Fotográfico")
                                ws_fotos['A1'] = "REPORTE FOTOGRÁFICO"
                                ws_fotos['A3'] = f"Proyecto: {row.get('Proyecto', '')}"
                                ws_fotos['A4'] = f"Protocolo N°: {row.get('Correlativo', '')}"
                                start_rows = [7, 25, 43]
                                for idx, i in enumerate((1,2,3)):
                                    foto_path = row.get(f"Foto{i}")
                                    if isinstance(foto_path, str) and foto_path and os.path.exists(foto_path):
                                        try:
                                            resized_path = resize_image_for_excel(foto_path)
                                            img = OpenpyxlImage(resized_path)
                                            ws_fotos.add_image(img, f"B{start_rows[idx]}")
                                        except Exception as e:
                                            st.warning(f"No se pudo cargar Foto{i}: {e}")
                                            
                            if "1" in wb.sheetnames: del wb["1"]
                            output = io.BytesIO(); wb.save(output); output.seek(0)
                            
                            st.success(f"✅ ¡Plantilla para {registro_a_generar} lista!")
                            st.download_button("⬇️ Descargar Excel Oficial", data=output, file_name=f"{registro_a_generar}_{row['Tag']}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
                        except Exception as e:
                            st.error(f"Error procesando el Excel: {e}")
                            
                elif row["Tipo_Protocolo"] == "PTC-MRP":
                    TEMPLATE_MRP = "LP16255BE-0138-0430-PTC-MRP_Resistencia de Pozo o Malla.xlsx"
                    if not os.path.exists(TEMPLATE_MRP):
                        st.error("❌ No se encontró la plantilla PTC-MRP.")
                    else:
                        try:
                            wb = openpyxl.load_workbook(TEMPLATE_MRP)
                            ws = wb.active
                            
                            def set_value(sheet, cell_coord, val):
                                cell = sheet[cell_coord]
                                if type(cell).__name__ == 'MergedCell':
                                    for merged_range in sheet.merged_cells.ranges:
                                        if cell_coord in merged_range:
                                            top_left = merged_range.coord.split(':')[0]
                                            sheet[top_left] = val
                                            return
                                else:
                                    cell.value = val

                            # Mapeo de Datos Generales
                            set_value(ws, 'Q5', row.get("Correlativo", ""))
                            set_value(ws, 'E4', row.get("Proyecto", ""))
                            set_value(ws, 'E5', row.get("Cliente", ""))
                            set_value(ws, 'E6', row.get("Lugar", ""))
                            set_value(ws, 'Q6', row.get("Contratista", ""))
                            set_value(ws, 'E7', row.get("Fecha", ""))
                            set_value(ws, 'Q7', row.get("Contrato", ""))
                            set_value(ws, 'E8', row.get("Plano", ""))
                            set_value(ws, 'Q8', row.get("Facility", ""))
                            set_value(ws, 'E9', row.get("Tag", ""))
                            set_value(ws, 'Q9', row.get("Disciplina", ""))
                            set_value(ws, 'E10', row.get("Sistema", ""))
                            set_value(ws, 'Q10', row.get("SubSistema", ""))
                            set_value(ws, 'E11', row.get("Estructura", ""))
                            set_value(ws, 'Q11', row.get("Hoja", ""))

                            # Mapeo de Equipo
                            set_value(ws, 'A14', row.get("Equipo_Prueba", ""))
                            set_value(ws, 'G14', row.get("Modelo", ""))
                            set_value(ws, 'M14', row.get("N_Serie", ""))
                            set_value(ws, 'T14', row.get("Fecha_Calib", ""))

                            # Mapeo de Datos Ambientales
                            set_value(ws, 'F18', row.get("Temperatura", ""))
                            set_value(ws, 'R18', row.get("Humedad", ""))
                            set_value(ws, 'F19', row.get("Lugar_Prueba", ""))
                            set_value(ws, 'R19', row.get("Hora_Prueba", ""))

                            # Mapeo de Mediciones
                            set_value(ws, 'B23', "X")
                            
                            dist_r3_val = float(row.get("Dist_R3", 14.5))
                            calibre_val = row.get("Calibre", "Varilla 3/4\"")
                            tipo_val = row.get("Tipo_Cond", "Electrolítica")
                            
                            # Row 26 (42%)
                            set_value(ws, 'B26', dist_r3_val)
                            set_value(ws, 'C26', round(dist_r3_val * 0.42, 1))
                            set_value(ws, 'D26', 0.42)
                            set_value(ws, 'E26', calibre_val)
                            set_value(ws, 'G26', tipo_val)
                            set_value(ws, 'I26', row.get("Val_42", ""))
                            
                            # Row 27 (52%)
                            set_value(ws, 'B27', dist_r3_val)
                            set_value(ws, 'C27', round(dist_r3_val * 0.52, 1))
                            set_value(ws, 'D27', 0.52)
                            set_value(ws, 'E27', calibre_val)
                            set_value(ws, 'G27', tipo_val)
                            set_value(ws, 'I27', row.get("Val_52", ""))
                            
                            # Row 28 (62%)
                            set_value(ws, 'B28', dist_r3_val)
                            set_value(ws, 'C28', round(dist_r3_val * 0.62, 1))
                            set_value(ws, 'D28', 0.62)
                            set_value(ws, 'E28', calibre_val)
                            set_value(ws, 'G28', tipo_val)
                            set_value(ws, 'I28', row.get("Val_62", ""))
                            
                            # Row 29 (72%)
                            set_value(ws, 'B29', dist_r3_val)
                            set_value(ws, 'C29', round(dist_r3_val * 0.72, 1))
                            set_value(ws, 'D29', 0.72)
                            set_value(ws, 'E29', calibre_val)
                            set_value(ws, 'G29', tipo_val)
                            set_value(ws, 'I29', row.get("Val_72", ""))
                            
                            # Row 30 (82%)
                            set_value(ws, 'B30', dist_r3_val)
                            set_value(ws, 'C30', round(dist_r3_val * 0.82, 1))
                            set_value(ws, 'D30', 0.82)
                            set_value(ws, 'E30', calibre_val)
                            set_value(ws, 'G30', tipo_val)
                            set_value(ws, 'I30', row.get("Val_82", ""))
                            
                            # RT (valor central 62%)
                            set_value(ws, 'O29', row.get("Val_62", ""))
                            
                            # Comentarios
                            set_value(ws, 'A60', row.get("Comentarios", ""))
                            
                            # --- SECCIÓN DE FOTOS ---
                            def resize_image_for_excel(img_path, max_width=400, max_height=300):
                                with PILImage.open(img_path) as img:
                                    img.thumbnail((max_width, max_height))
                                    if img.mode in ("RGBA", "P"):
                                        img = img.convert("RGB")
                                    img_byte_arr = io.BytesIO()
                                    img.save(img_byte_arr, format="JPEG")
                                    img_byte_arr.seek(0)
                                    return img_byte_arr

                            has_photos = any(isinstance(row.get(f"Foto{i}"), str) and row.get(f"Foto{i}") for i in (1,2,3))
                            if has_photos:
                                ws_fotos = wb.create_sheet(title="Reporte Fotográfico")
                                ws_fotos['A1'] = "REPORTE FOTOGRÁFICO"
                                ws_fotos['A3'] = f"Proyecto: {row.get('Proyecto', '')}"
                                ws_fotos['A4'] = f"Protocolo N°: {row.get('Correlativo', '')}"
                                
                                start_rows = [7, 25, 43]
                                for idx, i in enumerate((1,2,3)):
                                    foto_path = row.get(f"Foto{i}")
                                    if isinstance(foto_path, str) and foto_path and os.path.exists(foto_path):
                                        try:
                                            resized_path = resize_image_for_excel(foto_path)
                                            img = OpenpyxlImage(resized_path)
                                            ws_fotos.add_image(img, f"B{start_rows[idx]}")
                                        except Exception as e:
                                            st.warning(f"No se pudo cargar Foto{i}: {e}")
                            # --- FIN SECCIÓN FOTOS ---
                            
                            if "1" in wb.sheetnames:
                                del wb["1"]
                                
                            output = io.BytesIO()
                            wb.save(output)
                            output.seek(0)
                            
                            st.success(f"✅ ¡Plantilla para {registro_a_generar} lista!")
                            st.download_button(
                                label="⬇️ Descargar Excel Oficial",
                                data=output,
                                file_name=f"{registro_a_generar}_{row.get('Tag', 'MRP')}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                use_container_width=True
                            )
                        except Exception as e:
                            st.error(f"Error procesando el Excel: {e}")
                            
        with col_del:
            st.markdown("### 🗑️ Borrar Registro")
            with st.expander("Modo Administrador"):
                admin_pass = st.text_input("Contraseña Admin:", type="password")
                if admin_pass == st.secrets.get("admin_password", "ccecc2026"):
                    if not df_mostrar.empty:
                        opciones_borrar = ["(Ninguno)"] + df_mostrar["Correlativo"].iloc[::-1].tolist()
                    else:
                        opciones_borrar = ["(Ninguno)"]
                    registro_a_borrar = st.selectbox("Seleccione el correlativo a eliminar:", opciones_borrar, key="del")
                    
                    if st.button("🗑️ Eliminar Permanente"):
                        if registro_a_borrar != "(Ninguno)":
                            df = df[df["Correlativo"] != registro_a_borrar]
                            df.to_excel(DB_FILE, index=False)
                            st.success(f"Registro {registro_a_borrar} eliminado exitosamente.")
                            st.rerun()
                        else:
                            st.warning("Selecciona un registro válido para borrar.")
                elif admin_pass != "":
                    st.error("Contraseña incorrecta.")
    else:
        st.info("La base de datos está vacía. Registre un protocolo primero en la otra pestaña.")

with tab3:
    st.markdown("## 📊 Panel de Control (Gerencia)")
    st.markdown("Vista general del progreso de los protocolos y calidad (QA/QC).")
    
    FTPR_OBJETIVO = 90.0

    # Mapeo de checks
    MAPEO_CHECKS = {
        ("PTC-IPT", "Check1"): "1. Verificación Topográfica de ubicación",
        ("PTC-IPT", "Check2"): "2. Verificación de fundación de tratamiento",
        ("PTC-IPT", "Check3"): "3. Tratamiento de puesta a tierra",
        ("PTC-IPT", "Check4"): "4. Verificación de relleno y compactación",
        ("PTC-IPT", "Check5"): "5. Medición de resistencia puesta a tierra",
        ("PTC-IPT", "Check6"): "6. Ubicación e instalación de caja de registro",
        ("PTC-IPT", "Check7"): "7. Conexión de ánodos de aterramiento",
        ("PTC-IPT", "Check8"): "8. Rotulación y señalización",
        
        ("PTC-MRP", "Check1"): "1. Verificación topográfica de ubicación",
        ("PTC-MRP", "Check2"): "2. Verificación de excavación de zanja",
        ("PTC-MRP", "Check3"): "3. Instalación de conductor",
        ("PTC-MRP", "Check4"): "4. Instalación de puestas a tierra",
        ("PTC-MRP", "Check5"): "5. Tratamiento de suelo",
        ("PTC-MRP", "Check6"): "6. Relleno y compactación",
        ("PTC-MRP", "Check7"): "7. Medición de resistencia",
        ("PTC-MRP", "Check8"): "8. Ubicación e instalación de caja",
        ("PTC-MRP", "Check9"): "9. Conexión de mallas revisadas",
        ("PTC-MRP", "Check10"): "10. Rotulación y señalización",
        
        ("PTC-IST", "Check1"): "1. Verificación de planos y manuales",
        ("PTC-IST", "Check2"): "2. Verificación de base y/o soporte",
        ("PTC-IST", "Check3"): "3. Nivelación, plomada y alineamiento",
        ("PTC-IST", "Check4"): "4. Ajuste de pernos de anclaje",
        ("PTC-IST", "Check5"): "5. Instalación de tableros",
        ("PTC-IST", "Check6"): "6. Limpieza interior y exterior",
        ("PTC-IST", "Check7"): "7. Verificación de pintura",
        ("PTC-IST", "Check8"): "8. Rotulación de tableros",
        ("PTC-IST", "Check9"): "9. Verificación de hermeticidad",
        ("PTC-IST", "Check10"): "10. Conexión de cables de fuerza",
        ("PTC-IST", "Check11"): "11. Verificación de sistema a tierra",
        ("PTC-IST", "Check12"): "12. Mediciones de aislamiento",
        
        ("PTC-CBT", "Check1"): "1. Verificación de ruta de cable",
        ("PTC-CBT", "Check2"): "2. Verificación de soporte de cable",
        ("PTC-CBT", "Check3"): "3. Verificación de aislamiento de cables",
        ("PTC-CBT", "Check4"): "4. Instalación de cables (tendido)",
        ("PTC-CBT", "Check5"): "5. Peinado y amarre de cables",
        ("PTC-CBT", "Check6"): "6. Verificación de radios de curvatura",
        ("PTC-CBT", "Check7"): "7. Identificación de cables",
        ("PTC-CBT", "Check8"): "8. Conexión de cables",
        ("PTC-CBT", "Check9"): "9. Ajuste (torque) de conexiones",
        ("PTC-CBT", "Check10"): "10. Pruebas de continuidad",
        ("PTC-CBT", "Check11"): "11. Sellado de pases",
        
        ("PTC-CMT", "Check1"): "1. Verificación de ruta de cable",
        ("PTC-CMT", "Check2"): "2. Verificación de soporte",
        ("PTC-CMT", "Check3"): "3. Pruebas de aislamiento (Megado) previo",
        ("PTC-CMT", "Check4"): "4. Instalación de cables (tendido)",
        ("PTC-CMT", "Check5"): "5. Peinado y amarre de cables",
        ("PTC-CMT", "Check6"): "6. Verificación de radios de curvatura",
        ("PTC-CMT", "Check7"): "7. Identificación de cables",
        ("PTC-CMT", "Check8"): "8. Ejecución de terminaciones / empalmes",
        ("PTC-CMT", "Check9"): "9. Ajuste (torque) de conexiones",
        ("PTC-CMT", "Check10"): "10. Pruebas de aislamiento y VLF post",
        ("PTC-CMT", "Check11"): "11. Sellado de pases",
        
        ("PTC-IEI", "Check1"): "1. Verificación de ruta y soportes",
        ("PTC-IEI", "Check2"): "2. Verificación de equipos de iluminación",
        ("PTC-IEI", "Check3"): "3. Montaje de luminarias",
        ("PTC-IEI", "Check4"): "4. Instalación de tomacorrientes",
        ("PTC-IEI", "Check5"): "5. Instalación de interruptores",
        ("PTC-IEI", "Check6"): "6. Tendido y conexionado de cables",
        ("PTC-IEI", "Check7"): "7. Verificación de tableros de distribución",
        ("PTC-IEI", "Check8"): "8. Pruebas de aislamiento y continuidad",
        ("PTC-IEI", "Check9"): "9. Pruebas de funcionamiento",
        ("PTC-IEI", "Check10"): "10. Niveles de iluminación (Luxometría)",
        ("PTC-IEI", "Check11"): "11. Identificación y rotulado",

        ("PTC-BDE", "Check1"): "1. Verificación de ruta y soportes",
        ("PTC-BDE", "Check2"): "2. Montaje de bandejas portacables",
        ("PTC-BDE", "Check3"): "3. Uniones y accesorios de bandejas",
        ("PTC-BDE", "Check4"): "4. Alineamiento y nivelación",
        ("PTC-BDE", "Check5"): "5. Separación entre bandejas",
        ("PTC-BDE", "Check6"): "6. Puesta a tierra de bandejas",
        ("PTC-BDE", "Check7"): "7. Limpieza de bordes filosos",
        ("PTC-BDE", "Check8"): "8. Rotulado e identificación",
    }

    df_dash = load_db()
    if not df_dash.empty:
        df_dash['Fecha_Dt'] = pd.to_datetime(df_dash['Fecha'], format='%d/%m/%Y', errors='coerce')
        min_date = df_dash['Fecha_Dt'].min().date() if pd.notnull(df_dash['Fecha_Dt'].min()) else datetime.today().date()
        max_date = df_dash['Fecha_Dt'].max().date() if pd.notnull(df_dash['Fecha_Dt'].max()) else datetime.today().date()
        
        st.markdown("### 🔍 Filtros de Tablero")
        col_f1, col_f2, col_f3 = st.columns(3)
        with col_f1:
            fechas_seleccionadas = st.date_input("Rango de Fechas", [min_date, max_date], min_value=min_date, max_value=max_date)
        with col_f2:
            tipos_disponibles = df_dash["Tipo_Protocolo"].dropna().unique().tolist()
            tipos_sel = st.multiselect("Tipo de Protocolo", tipos_disponibles, default=tipos_disponibles)
        with col_f3:
            contratistas_disp = df_dash["Contratista"].dropna().unique().tolist()
            contratistas_sel = st.multiselect("Contratista", contratistas_disp, default=contratistas_disp)

        # Aplicar Filtros
        if len(fechas_seleccionadas) == 2:
            start_date, end_date = fechas_seleccionadas
            mask = (df_dash['Fecha_Dt'].dt.date >= start_date) & (df_dash['Fecha_Dt'].dt.date <= end_date)
            df_dash = df_dash.loc[mask]
        
        if not df_dash.empty:
            df_dash = df_dash[df_dash["Tipo_Protocolo"].isin(tipos_sel) & df_dash["Contratista"].isin(contratistas_sel)]

        if df_dash.empty:
            st.warning("No hay registros que coincidan con los filtros seleccionados.")
        else:
            # Cálculos KPI
            total_protocolos = len(df_dash)
            check_cols = [c for c in df_dash.columns if c.startswith("Check")]
            df_dash["Tiene_NC"] = df_dash[check_cols].apply(lambda row: any(str(val) == "No Conforme (NC)" for val in row), axis=1)
            filas_nc = df_dash["Tiene_NC"].sum()
            ftpr_pct = ((total_protocolos - filas_nc) / total_protocolos) * 100 if total_protocolos > 0 else 100.0
            
            delta_days = (end_date - start_date).days + 1 if len(fechas_seleccionadas) == 2 else 1
            promedio_diario = total_protocolos / delta_days if delta_days > 0 else total_protocolos

            col_m1, col_m2, col_m3, col_m4 = st.columns(4)
            col_m1.metric("Total Protocolos", total_protocolos)
            
            if ftpr_pct >= FTPR_OBJETIVO:
                col_m2.success(f"FTPR: {ftpr_pct:.1f}% (Objetivo: {FTPR_OBJETIVO}%)")
            elif ftpr_pct >= FTPR_OBJETIVO - 10:
                col_m2.warning(f"FTPR: {ftpr_pct:.1f}% (Objetivo: {FTPR_OBJETIVO}%)")
            else:
                col_m2.error(f"FTPR: {ftpr_pct:.1f}% (Objetivo: {FTPR_OBJETIVO}%)")
                
            col_m3.metric("Protocolos con NC", filas_nc)
            col_m4.metric("Promedio Diario", f"{promedio_diario:.1f}")

            st.markdown("---")

            # Desglose y Tendencia
            col_c1, col_c2 = st.columns([1, 2])
            with col_c1:
                st.markdown("### Desglose por Tipo")
                conteo_tipo = df_dash["Tipo_Protocolo"].value_counts().reset_index()
                conteo_tipo.columns = ["Tipo", "Cantidad"]
                st.bar_chart(conteo_tipo, x="Tipo", y="Cantidad")
            
            with col_c2:
                st.markdown("### Tendencia Temporal")
                agrupacion = st.radio("Granularidad:", ["Diario", "Semanal", "Mensual"], horizontal=True)
                if agrupacion == "Diario":
                    df_dash["Periodo"] = df_dash["Fecha_Dt"].dt.date
                elif agrupacion == "Semanal":
                    df_dash["Periodo"] = df_dash["Fecha_Dt"].dt.to_period("W").apply(lambda r: r.start_time.date())
                else:
                    df_dash["Periodo"] = df_dash["Fecha_Dt"].dt.to_period("M").apply(lambda r: r.start_time.date())
                
                conteo_fecha = df_dash["Periodo"].value_counts().sort_index().reset_index()
                conteo_fecha.columns = ["Periodo", "Cantidad"]
                conteo_fecha["Periodo"] = conteo_fecha["Periodo"].astype(str)
                st.line_chart(conteo_fecha, x="Periodo", y="Cantidad")

            st.markdown("---")
            
            # Top Defectos y Ranking
            col_r1, col_r2 = st.columns([1, 1])
            with col_r1:
                st.markdown("### Top Defectos (NC)")
                defectos = []
                for _, row in df_dash.iterrows():
                    tipo = row["Tipo_Protocolo"]
                    for col in check_cols:
                        if str(row.get(col, "")) == "No Conforme (NC)":
                            label = MAPEO_CHECKS.get((tipo, col), col)
                            label = label[:40] + "..." if len(label) > 40 else label
                            defectos.append(label)
                if defectos:
                    df_defectos = pd.Series(defectos).value_counts().reset_index()
                    df_defectos.columns = ["Defecto", "Cantidad"]
                    st.bar_chart(df_defectos, x="Defecto", y="Cantidad")
                else:
                    st.success("¡Excelente! No se registraron defectos.")

            with col_r2:
                st.markdown("### Ranking de Contratistas")
                ranking = []
                for contratista, group in df_dash.groupby("Contratista"):
                    tot = len(group)
                    nc = group["Tiene_NC"].sum()
                    ftpr = ((tot - nc) / tot) * 100
                    ranking.append({"Contratista": contratista, "Total Protocolos": tot, "% FTPR": ftpr})
                df_ranking = pd.DataFrame(ranking).sort_values("% FTPR", ascending=False)
                
                def color_ftpr(val):
                    if val >= FTPR_OBJETIVO: return 'background-color: #d4edda; color: black;'
                    elif val >= FTPR_OBJETIVO - 10: return 'background-color: #fff3cd; color: black;'
                    else: return 'background-color: #f8d7da; color: black;'
                
                st.dataframe(df_ranking.style.map(color_ftpr, subset=['% FTPR']), use_container_width=True, hide_index=True)

            st.markdown("---")
            
            # Tabla de Seguimiento de NC
            st.markdown("### Seguimiento de No Conformidades")
            df_nc = df_dash[df_dash["Tiene_NC"] == True]
            if not df_nc.empty:
                for _, row in df_nc.iterrows():
                    fallos = []
                    tipo = row["Tipo_Protocolo"]
                    for col in check_cols:
                        if str(row.get(col, "")) == "No Conforme (NC)":
                            fallos.append(MAPEO_CHECKS.get((tipo, col), col))
                    
                    with st.expander(f"🔴 {row.get('Correlativo', 'N/A')} | {row.get('Contratista', 'N/A')} | {row.get('Fecha', 'N/A')} | Tag: {row.get('Tag','N/A')}"):
                        st.markdown(f"**Checks Fallados:** {', '.join(fallos)}")
                        st.markdown(f"**Comentarios:** {row.get('Comentarios', 'Ninguno')}")
                        
                        fotos = [row.get("Foto1"), row.get("Foto2"), row.get("Foto3")]
                        fotos_validas = [f for f in fotos if pd.notnull(f) and os.path.exists(os.path.join("fotos_protocolos", str(f)))]
                        
                        if fotos_validas:
                            cols_img = st.columns(len(fotos_validas))
                            for idx, f in enumerate(fotos_validas):
                                cols_img[idx].image(os.path.join("fotos_protocolos", str(f)), width=200)
                        else:
                            st.info("No hay evidencia fotográfica disponible para este registro.")
            else:
                st.success("No hay No Conformidades en este rango.")
                
            # Exportar a Excel
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                pd.DataFrame([{
                    "Total Protocolos": total_protocolos,
                    "FTPR (%)": ftpr_pct,
                    "Protocolos con NC": filas_nc,
                    "Promedio Diario": promedio_diario
                }]).to_excel(writer, sheet_name="KPIs", index=False)
                
                conteo_tipo.to_excel(writer, sheet_name="Desglose por Tipo", index=False)
                df_ranking.to_excel(writer, sheet_name="Ranking Contratistas", index=False)
                
                if not df_nc.empty:
                    df_nc_export = df_nc[["Correlativo", "Fecha", "Contratista", "Tipo_Protocolo", "Tag", "Comentarios"]]
                    df_nc_export.to_excel(writer, sheet_name="Seguimiento NCs", index=False)
            
            output.seek(0)
            st.download_button(
                label="📥 Exportar Resumen de Dashboard a Excel",
                data=output,
                file_name=f"Dashboard_Gerencial_{datetime.now().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
    else:
        st.info("La base de datos está vacía. Registre un protocolo primero en la otra pestaña.")
